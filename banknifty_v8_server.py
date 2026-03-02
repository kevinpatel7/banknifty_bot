"""
BANK NIFTY v8 SERVER EDITION
Runs 24/7 on cloud server or Android phone (Termux)
Automatically starts/stops on market days only
Auto-restarts if it crashes
Logs everything to file
"""

import os, sys, signal, traceback
import logging
from logging.handlers import RotatingFileHandler

# ══════════════════════════════════════════════════════════════════════════════
# SERVER CONFIG
# ══════════════════════════════════════════════════════════════════════════════

# Paths — change if needed
LOG_FILE     = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "banknifty_server.log")
EXCEL_PATH   = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "banknifty_trades.xlsx")
PID_FILE     = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "banknifty.pid")

# ══════════════════════════════════════════════════════════════════════════════
# LOGGING SETUP
# ══════════════════════════════════════════════════════════════════════════════

# Rotate log file every 5MB, keep 3 backups
_handler = RotatingFileHandler(LOG_FILE, maxBytes=5*1024*1024, backupCount=3)
_handler.setFormatter(logging.Formatter(
    "%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
))
_console = logging.StreamHandler(sys.stdout)
_console.setFormatter(logging.Formatter("%(asctime)s %(message)s",
                                         datefmt="%H:%M:%S"))

logging.basicConfig(level=logging.INFO, handlers=[_handler, _console])
log = logging.getLogger("BankNifty")

# Override print so everything also goes to log file
_orig_print = print
def print(*args, **kwargs):
    msg = " ".join(str(a) for a in args)
    _orig_print(*args, **kwargs)
    log.info(msg)

# ══════════════════════════════════════════════════════════════════════════════
# MARKET HOURS CHECK
# ══════════════════════════════════════════════════════════════════════════════

import pytz as _pytz
_IST = _pytz.timezone("Asia/Kolkata")

def is_market_day():
    """Returns True if today is Mon-Fri (NSE trading day)."""
    return datetime.now(_IST).weekday() < 5   # 0=Mon, 4=Fri

def seconds_until_market():
    """How many seconds until 8:15 AM IST today or tomorrow."""
    from datetime import datetime as _dt, time as _t, timedelta as _td
    now  = _dt.now(_IST)
    open_today = now.replace(hour=8, minute=15, second=0, microsecond=0)
    if now < open_today:
        return (open_today - now).total_seconds()
    # Already past 8:15 — calculate for tomorrow
    tomorrow = now + _td(days=1)
    # Skip to next weekday
    while tomorrow.weekday() >= 5:
        tomorrow += _td(days=1)
    open_tomorrow = tomorrow.replace(hour=8, minute=15, second=0, microsecond=0)
    return (open_tomorrow - now).total_seconds()

def is_active_session():
    """Returns True between 8:00 AM and 4:00 PM IST on weekdays."""
    now = datetime.now(_IST)
    if now.weekday() >= 5:
        return False
    t = now.time()
    from datetime import time as _t
    return _t(8, 0) <= t <= _t(16, 0)

# ══════════════════════════════════════════════════════════════════════════════
# CRASH RECOVERY + AUTO-RESTART WRAPPER
# ══════════════════════════════════════════════════════════════════════════════

def _send_tg_startup(token, chat_id, mode):
    """Notify Telegram when server starts."""
    if not token or not chat_id:
        return
    try:
        import requests as _r
        msg = (
            "<b>✅ BankNifty Server Started</b>\n"
            f"Mode: {'PAPER TRADE' if mode else 'LIVE TRADE'}\n"
            f"Time: {datetime.now(_IST).strftime('%d %b %Y %H:%M IST')}\n"
            "System is running 24/7 automatically\n"
            "You will receive alerts throughout market hours"
        )
        _r.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            data={"chat_id": chat_id, "text": msg, "parse_mode": "HTML"},
            timeout=10
        )
    except:
        pass

def _send_tg_crash(token, chat_id, error):
    """Notify Telegram on crash."""
    if not token or not chat_id:
        return
    try:
        import requests as _r
        msg = (
            "<b>⚠️ BankNifty Server Crashed — Restarting</b>\n"
            f"Error: {str(error)[:200]}\n"
            "System will auto-restart in 60 seconds"
        )
        _r.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            data={"chat_id": chat_id, "text": msg, "parse_mode": "HTML"},
            timeout=10
        )
    except:
        pass

# Write PID file so we can monitor the process
with open(PID_FILE, "w") as _f:
    _f.write(str(os.getpid()))

log.info("="*60)
log.info("BANK NIFTY SERVER STARTING")
log.info(f"PID: {os.getpid()}")
log.info(f"Log: {LOG_FILE}")
log.info(f"Excel: {EXCEL_PATH}")
log.info("="*60)

# Smart sleep — wait for market hours if starting too early
_now = datetime.now(_IST) if False else __import__("datetime").datetime.now(
    __import__("pytz").timezone("Asia/Kolkata"))
_t_now = _now.time()
_market_start = __import__("datetime").time(8, 15)
_market_end   = __import__("datetime").time(16, 0)

# ── INSTANT TELEGRAM PING — fires before any sleeping ────────────────────
try:
    import requests as _rq_early
    _token_early = _os.environ.get("TELEGRAM_TOKEN","7753587413:AAHWLS-qZ7aVxMCycHiT8yhaUcEHNlRg_hU")
    _chat_early  = _os.environ.get("TELEGRAM_CHAT_ID","8129647943")
    _now_str_early = __import__("datetime").datetime.now(
        __import__("pytz").timezone("Asia/Kolkata")).strftime("%d %b %Y %H:%M IST")
    _ping_early = (
        "<b>\u2705 BankNifty Bot is ONLINE</b>\n"
        f"Started: {_now_str_early}\n"
        "Mode: PAPER TRADE\n"
        "\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\n"
        "Bot will sleep until 8:15 AM IST\n"
        "Then train ML and send morning briefing\n"
        "Trade signals follow automatically"
    )
    _rq_early.post(
        f"https://api.telegram.org/bot{_token_early}/sendMessage",
        data={"chat_id": _chat_early, "text": _ping_early, "parse_mode": "HTML"},
        timeout=10
    )
    print("STARTUP TELEGRAM SENT!")
except Exception as _ep:
    print(f"Early ping failed: {_ep}")

if _now.weekday() >= 5:
    _secs = seconds_until_market()
    _hrs  = int(_secs//3600)
    _mins = int((_secs%3600)//60)
    log.info(f"Weekend — sleeping {_hrs}h {_mins}m until next market day")
    print(f"Weekend detected. Sleeping until Monday 8:15 AM IST...")
    import time as _t_mod
    _t_mod.sleep(min(_secs, 3600))   # wake up every hour to check
elif _t_now < _market_start:
    _secs = seconds_until_market()
    _mins = int(_secs//60)
    log.info(f"Pre-market — sleeping {_mins} min until 8:15 AM")
    print(f"Pre-market. Sleeping {_mins} minutes until 8:15 AM IST...")
    import time as _t_mod
    _t_mod.sleep(max(0, _secs - 60))   # wake up 1 min before market

import subprocess as _subp, sys
print("Installing/checking libraries...")
_pkgs = ["yfinance","pandas","numpy","scikit-learn","pytz",
         "lightgbm","xgboost","hmmlearn","tabulate",
         "requests","beautifulsoup4","lxml","openpyxl"]
for _p in _pkgs:
    try:
        __import__(_p.replace("-","_").split("[")[0])
    except ImportError:
        _subp.check_call([sys.executable,"-m","pip","install","-q",_p,
                         "--break-system-packages"],
                        stdout=_subp.DEVNULL, stderr=_subp.DEVNULL)
print("All libraries ready!")

import yfinance as yf
import pandas as pd
import numpy as np
import requests, time, warnings, json, re
import lightgbm as lgb
import xgboost as xgb
from hmmlearn import hmm
from bs4 import BeautifulSoup
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import TimeSeriesSplit
from sklearn.metrics import accuracy_score
from tabulate import tabulate
from datetime import datetime, time as dt_time, timedelta
import pytz

warnings.filterwarnings("ignore")
IST = pytz.timezone("Asia/Kolkata")

# ══════════════════════════════════════════════════════════════════════════════
# ★ USER SETTINGS — EDIT THESE ★
# ══════════════════════════════════════════════════════════════════════════════

CAPITAL          = 50000     # ← Your trading capital in ₹
RISK_PER_TRADE   = 1.0       # ← BEGINNER SETTING: only 1% = ₹500 max loss per trade
MAX_TRADES_DAY   = 2         # ← BEGINNER SETTING: max 2 trades per day only
MIN_RR           = 2.0       # ← BEGINNER SETTING: only trade if reward is 2x the risk

# Precision filters — raised for beginner (higher bar = fewer but better signals)
MIN_CONFLUENCE   = 68        # ← raised from 62: only very strong signals
MIN_ADX          = 20        # ← raised from 18: only clear trending markets
MIN_CONFIDENCE   = 65        # ← raised from 60: only high-confidence ML signals

# BEGINNER MODE — extra safety rules
BEGINNER_MODE         = True   # enables extra warnings and guidance
PAPER_TRADE_MODE      = True   # SET TO False ONLY when you're ready for real money
MAX_PREMIUM_PER_LOT   = 200    # ← never buy options above ₹200 premium (controls cost)
AVOID_EXPIRY_DAY      = True   # ← skip trading on Thursday (expiry — very risky)
CHECK_INTERVAL   = 120       # seconds between refreshes

# ══════════════════════════════════════════════════════════════════════════════
# ★ ALERT SETTINGS — fill these to get phone notifications ★
# ══════════════════════════════════════════════════════════════════════════════
# HOW TO SETUP (takes 5 min):
#
# TELEGRAM (free, recommended):
#   Step 1: Open Telegram → search @BotFather → send /newbot
#   Step 2: Give it a name e.g. "BankNiftyAlerts"
#   Step 3: BotFather gives you a TOKEN — paste below
#   Step 4: Search @userinfobot on Telegram → it shows your CHAT_ID — paste below
#   Step 5: Done! You will get instant messages on Telegram
#
# FAST2SMS (actual phone SMS — free 100 SMS on signup):
#   Step 1: Go to fast2sms.com → sign up free
#   Step 2: Go to Dev API → copy your API key → paste below
#   Step 3: Enter your 10-digit mobile number below
#
# Leave blank ("") if you don't want that alert type

import os as _os
TELEGRAM_TOKEN   = _os.environ.get("TELEGRAM_TOKEN","7753587413:AAHWLS-qZ7aVxMCycHiT8yhaUcEHNlRg_hU")
TELEGRAM_CHAT_ID = _os.environ.get("TELEGRAM_CHAT_ID","8129647943")

# ── INSTANT STARTUP PING (fires before anything else) ─────────────────────
try:
    import requests as _rq_ping
    _ping_msg = (
        "<b>\u2705 BankNifty Bot Starting...</b>\n"
        "GitHub server is online\n"
        "Installing libraries and training ML...\n"
        "This takes 3-5 minutes. Stand by."
    )
    _rq_ping.post(
        f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
        data={"chat_id": TELEGRAM_CHAT_ID, "text": _ping_msg,
              "parse_mode": "HTML"},
        timeout=10
    )
    print("Startup ping sent to Telegram!")
except Exception as _pe:
    print(f"Startup ping failed: {_pe}")

FAST2SMS_API_KEY = ""    # ← paste your Fast2SMS API key here
YOUR_MOBILE      = ""    # ← your 10-digit Indian mobile number (no +91)

# Alert preferences
ALERT_ON_SIGNAL     = True   # send alert when BUY/SELL signal fires
ALERT_ON_TARGET     = True   # send alert when price hits T1/T2/T3 levels
ALERT_ON_SL         = True   # send alert when stop loss level is hit
ALERT_MORNING_BRIEF = True   # send morning briefing summary to phone
ALERT_FORCE_EXIT    = True   # send 3:15 PM exit reminder



# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
TICKER          = "^NSEBANK"
LOT_SIZE        = 25
STRIKE_INTERVAL = 100
INTRADAY_TF     = "15m"

MARKET_OPEN  = dt_time(9, 15);   MARKET_CLOSE = dt_time(15, 30)
OR_END       = dt_time(9, 45);   LUNCH_START  = dt_time(12, 30)
LUNCH_END    = dt_time(13, 30);  POWER_HOUR   = dt_time(14, 30)
FORCE_EXIT   = dt_time(15, 15)   # mandatory close all positions

# Bank Nifty constituents with weights
BANKS = {
    "HDFCBANK.NS"  : ("HDFC Bank",           29.5),
    "ICICIBANK.NS" : ("ICICI Bank",           23.1),
    "KOTAKBANK.NS" : ("Kotak Mahindra Bank",  12.8),
    "AXISBANK.NS"  : ("Axis Bank",             8.9),
    "SBIN.NS"      : ("State Bank of India",   7.6),
    "INDUSINDBK.NS": ("IndusInd Bank",          5.2),
    "BANDHANBNK.NS": ("Bandhan Bank",           2.8),
    "FEDERALBNK.NS": ("Federal Bank",           2.4),
    "BANKBARODA.NS": ("Bank of Baroda",         2.1),
    "IDFCFIRSTB.NS": ("IDFC First Bank",        2.0),
    "AUBANK.NS"    : ("AU Small Finance Bank",  1.9),
    "PNB.NS"       : ("Punjab National Bank",   1.7),
}
HEAVYWEIGHTS = {"HDFCBANK.NS", "ICICIBANK.NS", "KOTAKBANK.NS"}

GLOBAL_TICKERS = {
    "^GSPC": "S&P 500", "^IXIC": "Nasdaq", "^DJI": "Dow Jones",
    "^HSI" : "Hang Seng", "GC=F": "Gold", "CL=F": "Crude Oil",
    "^N225": "Nikkei",
}

# Keywords that affect Bank Nifty
BULLISH_KEYWORDS = [
    "rate cut", "rbi cut", "repo cut", "stimulus", "gdp growth",
    "credit growth", "strong results", "beats estimates", "profit rises",
    "npa falls", "capital adequacy", "upgrade", "buy rating",
    "fii buying", "net buyer", "record high", "bullish"
]
BEARISH_KEYWORDS = [
    "rate hike", "rbi hike", "inflation", "npa rises", "bad loans",
    "fraud", "rbi penalty", "downgrade", "sell rating", "fii selling",
    "net seller", "loss", "misses estimates", "weak results",
    "liquidity crisis", "default", "bearish", "recession"
]

print("=" * 80)
print("🏦  BANK NIFTY ULTIMATE DAY TRADING GUIDE v6.0")
print("=" * 80)
print(f"  Capital       : ₹{CAPITAL:,}")
print(f"  Risk/trade    : {RISK_PER_TRADE}% = ₹{round(CAPITAL*RISK_PER_TRADE/100):,} max loss")
print(f"  Max trades    : {MAX_TRADES_DAY}/day")
print(f"  Min R:R       : 1:{MIN_RR}")
print(f"  ML engine     : LightGBM + XGBoost + LR | HMM regime\n")

# ══════════════════════════════════════════════════════════════════════════════
# DAILY TRADE LOG
# ══════════════════════════════════════════════════════════════════════════════

class DayLog:
    """Tracks every signal and trade for the day."""
    def __init__(self):
        self.reset()

    def reset(self):
        self.signals      = []    # all signals given today
        self.trades       = []    # trades actually taken (manual tracking)
        self.trades_taken = 0
        self.morning_bias = None
        self.or_high      = None
        self.or_low       = None
        self.or_range     = None
        self.day_plan     = []
        self.date         = None
        self.first_signal  = None
        self.active_trade  = None   # currently open trade being tracked
        self.excel_signal_row = None  # Excel row for current trade

    def log_signal(self, time_, direction, score, confidence, spot, strike, typ, ep, sl, t1, t2, t3):
        entry = {
            "time": time_, "direction": direction, "score": score,
            "confidence": confidence, "spot": spot, "strike": strike,
            "type": typ, "entry": ep, "sl": sl,
            "t1": t1, "t2": t2, "t3": t3,
        }
        self.signals.append(entry)
        if self.first_signal is None:
            self.first_signal = entry

    def summary(self):
        print(f"\n{'═'*80}")
        print(f"📋  TODAY'S SIGNAL LOG  ({self.date})")
        print(f"{'═'*80}")
        if not self.signals:
            print("  No trade signals generated today.")
            return
        rows = []
        for s in self.signals:
            rows.append([
                s["time"], s["direction"], f'{s["score"]}/100',
                f'{s["confidence"]}%', f'₹{s["spot"]:,.0f}',
                f'{s["strike"]} {s["type"]}', f'₹{s["entry"]}',
                f'₹{s["sl"]}', f'₹{s["t1"]}'
            ])
        print(tabulate(rows,
            headers=["Time","Dir","Score","ML%","Spot","Contract","Entry","SL","T1"],
            tablefmt="simple"))
        print(f"\n  Total signals today: {len(self.signals)}")
        print(f"  Trades taken today : {self.trades_taken}/{MAX_TRADES_DAY}")

day_log = DayLog()

# ══════════════════════════════════════════════════════════════════════════════
# UTILITY
# ══════════════════════════════════════════════════════════════════════════════

def safe_s(df, col):
    s = df[col]
    return s.iloc[:, 0] if isinstance(s, pd.DataFrame) else s

def pct(a, b):
    return round(((a - b) / b) * 100, 2) if b and b != 0 else 0.0

def atm(price):
    return round(price / STRIKE_INTERVAL) * STRIKE_INTERVAL

def premium_est(spot, strike, typ, vix=15, days=3):
    intr = max(spot-strike, 0) if typ=="CE" else max(strike-spot, 0)
    extr = (spot * vix/100) * np.sqrt(days/365) * 0.4
    return round(intr + extr, 2)

def max_lots_from_risk(entry_prem, sl_prem):
    """Calculate max lots we can trade given risk per trade setting."""
    risk_per_lot = abs(entry_prem - sl_prem) * LOT_SIZE
    if risk_per_lot <= 0:
        return 1
    max_risk = CAPITAL * RISK_PER_TRADE / 100
    lots = int(max_risk / risk_per_lot)
    return max(1, min(lots, 5))   # cap at 5 lots for safety

def session_name(t):
    if   t < MARKET_OPEN:                  return "PRE_MARKET"
    elif t > MARKET_CLOSE:                 return "CLOSED"
    elif t < OR_END:                       return "OPENING_RANGE"
    elif LUNCH_START <= t <= LUNCH_END:    return "LUNCH"
    elif t >= FORCE_EXIT:                  return "FORCE_EXIT"
    elif t >= POWER_HOUR:                  return "POWER_HOUR"
    else:                                  return "REGULAR"

# ══════════════════════════════════════════════════════════════════════════════
# NEWS SENTIMENT SCRAPER  (NEW — free financial news)
# ══════════════════════════════════════════════════════════════════════════════

def fetch_news_sentiment():
    """     Scrapes financial news headlines from multiple free sources.     Returns sentiment score and key headlines affecting Bank Nifty.     """
    headlines = []
    sentiment_score = 0   # positive = bullish, negative = bearish

    HEADERS = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
    }

    sources = [
        # Economic Times markets
        ("https://economictimes.indiatimes.com/markets/stocks/news",
         "a", {"class": "eachStory"}, "text"),
        # Moneycontrol
        ("https://www.moneycontrol.com/news/business/markets/",
         "a", {}, "title"),
        # LiveMint
        ("https://www.livemint.com/market/stock-market-news",
         "h2", {}, "text"),
    ]

    for url, tag, attrs, mode in sources:
        try:
            resp = requests.get(url, headers=HEADERS, timeout=8)
            soup = BeautifulSoup(resp.text, "lxml")
            elements = soup.find_all(tag, attrs, limit=15)
            for el in elements:
                text = el.get_text(strip=True) if mode=="text" else el.get("title","")
                if len(text) > 20:
                    headlines.append(text[:150])
        except:
            pass

    # Score headlines
    bull_count = 0; bear_count = 0
    relevant_headlines = []

    BANK_KEYWORDS = ["bank", "nifty", "hdfc", "icici", "kotak", "sbi", "axis",
                     "rbi", "repo", "credit", "npa", "finance", "rate", "fii",
                     "banking", "indusind", "federal", "sensex", "market"]

    for h in headlines:
        hl = h.lower()
        # Only count if relevant to banking/markets
        if not any(kw in hl for kw in BANK_KEYWORDS):
            continue
        relevant_headlines.append(h)
        for kw in BULLISH_KEYWORDS:
            if kw in hl:
                bull_count += 1
                sentiment_score += 1
                break
        for kw in BEARISH_KEYWORDS:
            if kw in hl:
                bear_count += 1
                sentiment_score -= 1
                break

    # Normalize to -100 to +100
    total = bull_count + bear_count
    if total > 0:
        norm_score = round(((bull_count - bear_count) / total) * 100)
    else:
        norm_score = 0

    if norm_score >= 30:    bias = "🟢 BULLISH NEWS"
    elif norm_score >= 10:  bias = "🟢 MILDLY BULLISH"
    elif norm_score <= -30: bias = "🔴 BEARISH NEWS"
    elif norm_score <= -10: bias = "🔴 MILDLY BEARISH"
    else:                   bias = "⚪ NEUTRAL NEWS"

    return {
        "score"      : norm_score,
        "bias"       : bias,
        "bull_count" : bull_count,
        "bear_count" : bear_count,
        "headlines"  : relevant_headlines[:8],   # top 8 relevant
        "total_found": len(relevant_headlines),
    }


def fetch_rbi_events():
    """     Returns known high-impact RBI / economic events.     Checks if today or tomorrow has a scheduled event.     """
    # Key RBI MPC dates 2025-26 (update annually)
    RBI_MPC_DATES = [
        "2025-04-09", "2025-06-06", "2025-08-08",
        "2025-10-08", "2025-12-05", "2026-02-07",
        "2026-04-09", "2026-06-05",
    ]

    today_str = datetime.now(IST).strftime("%Y-%m-%d")
    tmrw_str  = (datetime.now(IST) + timedelta(days=1)).strftime("%Y-%m-%d")

    events = []
    for d in RBI_MPC_DATES:
        if d == today_str:
            events.append(("🚨 TODAY", "RBI MPC Policy Decision — EXTREME VOLATILITY EXPECTED"))
        elif d == tmrw_str:
            events.append(("⚠️  TOMORROW", "RBI MPC Policy Decision — High volatility risk"))

    # US Fed dates (affect global markets → Bank Nifty)
    FED_DATES = [
        "2025-03-19", "2025-05-07", "2025-06-18", "2025-07-30",
        "2025-09-17", "2025-10-29", "2025-12-10", "2026-01-28",
        "2026-03-18", "2026-04-29",
    ]
    for d in FED_DATES:
        if d == today_str:
            events.append(("🌐 TODAY", "US Federal Reserve Rate Decision — Global impact"))
        elif d == tmrw_str:
            events.append(("🌐 TOMORROW", "US Fed Decision — Global markets may be cautious"))

    return events

# ══════════════════════════════════════════════════════════════════════════════
# NSE LIVE OPTIONS CHAIN
# ══════════════════════════════════════════════════════════════════════════════

NSE_HEADERS = {
    "User-Agent"     : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/120.0.0.0 Safari/537.36",
    "Accept"         : "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer"        : "https://www.nseindia.com/",
    "Connection"     : "keep-alive",
}

_nse_sess = None; _nse_sess_ts = None

def get_nse_session():
    global _nse_sess, _nse_sess_ts
    now = datetime.now()
    if _nse_sess is None or (now - _nse_sess_ts).seconds > 1800:
        s = requests.Session()
        s.headers.update(NSE_HEADERS)
        try:
            s.get("https://www.nseindia.com", timeout=10); time.sleep(1)
            s.get("https://www.nseindia.com/option-chain", timeout=10); time.sleep(0.5)
        except: pass
        _nse_sess = s; _nse_sess_ts = now
    return _nse_sess

def fetch_options_chain():
    empty = {"success":False,"spot":0,"pcr":1.0,"pcr_signal":"⚪ N/A",
             "max_pain":0,"max_pain_dist":0,"expiry":"","total_ce_oi":0,
             "total_pe_oi":0,"strikes":[],"resistance":[],"support":[],
             "oi_trend":"NEUTRAL","pe_buildup":False,"ce_buildup":False}
    try:
        sess = get_nse_session()
        url  = "https://www.nseindia.com/api/option-chain-indices?symbol=BANKNIFTY"
        resp = sess.get(url, timeout=15)
        if resp.status_code != 200:
            return empty

        data    = resp.json()
        records = data.get("records", {})
        spot    = float(records.get("underlyingValue", 0))
        expiries= records.get("expiryDates", [])
        if not expiries or spot == 0:
            return empty

        expiry = expiries[0]
        sd = {}
        for item in records.get("data", []):
            if item.get("expiryDate") != expiry: continue
            sk = item.get("strikePrice", 0)
            if sk == 0: continue
            if sk not in sd:
                sd[sk] = {"strike":sk,"ce_oi":0,"ce_chg_oi":0,"ce_ltp":0,
                          "ce_iv":0,"pe_oi":0,"pe_chg_oi":0,"pe_ltp":0,"pe_iv":0}
            ce = item.get("CE", {}); pe = item.get("PE", {})
            if ce:
                sd[sk]["ce_oi"]     = ce.get("openInterest", 0)
                sd[sk]["ce_chg_oi"] = ce.get("changeinOpenInterest", 0)
                sd[sk]["ce_ltp"]    = ce.get("lastPrice", 0)
                sd[sk]["ce_iv"]     = ce.get("impliedVolatility", 0)
            if pe:
                sd[sk]["pe_oi"]     = pe.get("openInterest", 0)
                sd[sk]["pe_chg_oi"] = pe.get("changeinOpenInterest", 0)
                sd[sk]["pe_ltp"]    = pe.get("lastPrice", 0)
                sd[sk]["pe_iv"]     = pe.get("impliedVolatility", 0)

        sl = sorted(sd.values(), key=lambda x: x["strike"])

        # PCR
        tce = sum(s["ce_oi"] for s in sl); tpe = sum(s["pe_oi"] for s in sl)
        pcr = round(tpe/(tce+1e-9), 2)
        if pcr>=1.5:   pcr_sig="🟢 STRONG BULL (institutions hedging)"
        elif pcr>=1.2: pcr_sig="🟢 BULLISH (put-heavy)"
        elif pcr>=0.8: pcr_sig="⚪ NEUTRAL"
        elif pcr>=0.5: pcr_sig="🔴 BEARISH (call-heavy)"
        else:          pcr_sig="🔴 STRONG BEAR (extreme call buying)"

        # Max Pain
        mp_strike=0; min_loss=float("inf")
        for ts in sd.keys():
            loss = sum(s["ce_oi"]*max(0,s["strike"]-ts) +
                       s["pe_oi"]*max(0,ts-s["strike"]) for s in sl)
            if loss < min_loss: min_loss=loss; mp_strike=ts

        # Support / Resistance
        atm_ = atm(spot)
        above = sorted([s for s in sl if s["strike"]>spot], key=lambda x:x["ce_oi"], reverse=True)[:3]
        below = sorted([s for s in sl if s["strike"]<spot], key=lambda x:x["pe_oi"], reverse=True)[:3]

        # OI trend (near ATM)
        near = [s for s in sl if abs(s["strike"]-atm_)<=3*STRIKE_INTERVAL]
        net_ce = sum(s["ce_chg_oi"] for s in near)
        net_pe = sum(s["pe_chg_oi"] for s in near)
        if net_pe>0 and net_ce<0:   oi_trend="BULLISH"
        elif net_ce>0 and net_pe<0: oi_trend="BEARISH"
        else:                        oi_trend="NEUTRAL"

        atm_strikes = [s for s in sl if abs(s["strike"]-atm_)<=10*STRIKE_INTERVAL]

        return {"success":True,"spot":spot,"pcr":pcr,"pcr_signal":pcr_sig,
                "max_pain":mp_strike,"max_pain_dist":round(((mp_strike-spot)/spot)*100,2),
                "expiry":expiry,"total_ce_oi":tce,"total_pe_oi":tpe,
                "strikes":atm_strikes,"resistance":[s["strike"] for s in above],
                "support":[s["strike"] for s in below],"oi_trend":oi_trend,
                "pe_buildup":net_pe>0,"ce_buildup":net_ce>0}
    except Exception as e:
        print(f"  ⚠️  NSE error: {e}")
        return empty

def get_real_premium(strike, typ, oc):
    if not oc["success"]: return None
    for s in oc["strikes"]:
        if s["strike"]==strike:
            ltp = s["ce_ltp"] if typ=="CE" else s["pe_ltp"]
            return ltp if ltp>0 else None
    return None

# ══════════════════════════════════════════════════════════════════════════════
# INDICATORS
# ══════════════════════════════════════════════════════════════════════════════

def rsi_f(c, p=14):
    d=c.diff(); g=d.where(d>0,0).rolling(p).mean()
    l=(-d.where(d<0,0)).rolling(p).mean()
    return 100-(100/(1+g/(l+1e-9)))

def adx_f(h,l,c,p=14):
    tr=pd.concat([h-l,(h-c.shift()).abs(),(l-c.shift()).abs()],axis=1).max(axis=1)
    atr=tr.rolling(p).mean()
    up=h.diff(); dn=-l.diff()
    pdm=up.where((up>dn)&(up>0),0); ndm=dn.where((dn>up)&(dn>0),0)
    pdi=100*pdm.rolling(p).mean()/(atr+1e-9)
    ndi=100*ndm.rolling(p).mean()/(atr+1e-9)
    dx=100*(pdi-ndi).abs()/(pdi+ndi+1e-9)
    return dx.rolling(p).mean(), pdi, ndi

def supertrend_f(h,l,c,p=10,m=3):
    hl2=(h+l)/2
    tr=pd.concat([h-l,(h-c.shift()).abs(),(l-c.shift()).abs()],axis=1).max(axis=1)
    a=tr.rolling(p).mean(); up=hl2+m*a; dn=hl2-m*a
    st=pd.Series(np.nan,index=c.index); di=pd.Series(0,index=c.index)
    for i in range(1,len(c)):
        fu=min(up.iloc[i],up.iloc[i-1]) if c.iloc[i-1]<=up.iloc[i-1] else up.iloc[i]
        fl=max(dn.iloc[i],dn.iloc[i-1]) if c.iloc[i-1]>=dn.iloc[i-1] else dn.iloc[i]
        pst=st.iloc[i-1] if not np.isnan(st.iloc[i-1]) else fl
        d_=-1 if (pst==up.iloc[i-1] and c.iloc[i]>fu) or (pst!=up.iloc[i-1] and c.iloc[i]>=fl) else 1
        st.iloc[i]=fl if d_==-1 else fu; di.iloc[i]=d_
    return st, di

def stochrsi_f(c,rp=14,sp=14,sm=3):
    r=rsi_f(c,rp)
    k=100*(r-r.rolling(sp).min())/(r.rolling(sp).max()-r.rolling(sp).min()+1e-9)
    return k, k.rolling(sm).mean()

def bollinger_f(c,p=20,sd=2):
    mid=c.rolling(p).mean(); sig_=c.rolling(p).std()
    up=mid+sd*sig_; lo=mid-sd*sig_
    return mid,up,lo,(c-lo)/(up-lo+1e-9),(up-lo)/(mid+1e-9)

def cci_f(h,l,c,p=20):
    tp=(h+l+c)/3; mad=tp.rolling(p).apply(lambda x:np.abs(x-x.mean()).mean())
    return (tp-tp.rolling(p).mean())/(0.015*mad+1e-9)

def atr_f(h,l,c,p=14):
    tr=pd.concat([h-l,(h-c.shift()).abs(),(l-c.shift()).abs()],axis=1).max(axis=1)
    return tr.rolling(p).mean()

def vwap_f(df):
    h=safe_s(df,"High"); l=safe_s(df,"Low"); c=safe_s(df,"Close"); v=safe_s(df,"Volume")
    tp=(h+l+c)/3
    return float((tp*v).sum()/v.sum()) if v.sum()>0 else float(c.iloc[-1])

# ══════════════════════════════════════════════════════════════════════════════
# FEATURE ENGINEERING  (65+ features)
# ══════════════════════════════════════════════════════════════════════════════

def build_features(df, news_score=0):
    data = pd.DataFrame(index=df.index)
    for col in ["Open","High","Low","Close","Volume"]:
        if col in df.columns: data[col]=safe_s(df,col)

    c=data["Close"]; h=data["High"]; l=data["Low"]; v=data["Volume"]

    # VIX
    try:
        vx=yf.download("^INDIAVIX",period="2y",auto_adjust=True,progress=False)["Close"]
        if isinstance(vx,pd.DataFrame): vx=vx.iloc[:,0]
        data["VIX"]=vx.reindex(data.index,method="ffill")
        data["VIX_change"]=data["VIX"].pct_change(1)*100
        data["VIX_vs_avg"]=data["VIX"]-data["VIX"].rolling(5).mean()
    except:
        data["VIX"]=15.0; data["VIX_change"]=0.0; data["VIX_vs_avg"]=0.0

    # Global markets
    for tkr,nm in [("^GSPC","SP500"),("^IXIC","Nasdaq"),("CL=F","Oil"),("GC=F","Gold"),("^N225","Nikkei")]:
        try:
            g=yf.download(tkr,period="2y",auto_adjust=True,progress=False)["Close"]
            if isinstance(g,pd.DataFrame): g=g.iloc[:,0]
            data[f"{nm}_ret"]=g.pct_change(1).reindex(data.index,method="ffill")*100
            data[f"{nm}_3d"]=g.pct_change(3).reindex(data.index,method="ffill")*100
        except:
            data[f"{nm}_ret"]=0.0; data[f"{nm}_3d"]=0.0

    # Returns
    for p in [1,2,3,5,10,20]: data[f"Ret_{p}d"]=c.pct_change(p)

    # Moving averages
    for p in [5,10,20,50,100,200]: data[f"MA_{p}"]=c.rolling(p).mean()
    data["EMA_9"]=c.ewm(span=9).mean(); data["EMA_21"]=c.ewm(span=21).mean()
    data["EMA_50"]=c.ewm(span=50).mean()

    # Slopes
    data["MA5_slp"] =data["MA_5"].diff(3)/(data["MA_5"].shift(3)+1e-9)
    data["MA20_slp"]=data["MA_20"].diff(5)/(data["MA_20"].shift(5)+1e-9)
    data["MA50_slp"]=data["MA_50"].diff(10)/(data["MA_50"].shift(10)+1e-9)

    # Price vs MAs
    data["P_MA20"] =(c-data["MA_20"])/(data["MA_20"]+1e-9)
    data["P_MA50"] =(c-data["MA_50"])/(data["MA_50"]+1e-9)
    data["P_MA200"]=(c-data["MA_200"])/(data["MA_200"]+1e-9)
    data["MA50_200"]=(data["MA_50"]-data["MA_200"])/(data["MA_200"]+1e-9)
    data["EMA9_21"]=(data["EMA_9"]-data["EMA_21"])/(data["EMA_21"]+1e-9)

    # MACD
    data["MACD"]     =c.ewm(span=12).mean()-c.ewm(span=26).mean()
    data["MACD_sig"] =data["MACD"].ewm(span=9).mean()
    data["MACD_hist"]=data["MACD"]-data["MACD_sig"]
    data["MACD_slp"] =data["MACD_hist"].diff(2)

    # RSI
    data["RSI"]    =rsi_f(c); data["RSI_slp"]=data["RSI"].diff(3)
    data["RSI_9"]  =rsi_f(c,9)
    sk,sd_=stochrsi_f(c)
    data["Stoch_K"]=sk; data["Stoch_D"]=sd_; data["Stoch_diff"]=sk-sd_

    # ADX + Supertrend
    a,pdi,ndi=adx_f(h,l,c)
    data["ADX"]=a; data["DI_diff"]=pdi-ndi; data["PDI"]=pdi; data["NDI"]=ndi
    st,std=supertrend_f(h,l,c)
    data["ST_dir"]=std; data["P_vs_ST"]=(c-st)/(st+1e-9)

    # Bollinger
    _,_,_,pctb,bbw=bollinger_f(c)
    data["BB_pctb"]=pctb; data["BB_width"]=bbw
    data["BB_squeeze"]=(bbw<bbw.rolling(50).mean()).astype(int)

    # CCI
    data["CCI"]=cci_f(h,l,c)

    # ATR & Volatility
    data["ATR"]    =atr_f(h,l,c); data["ATR_pct"]=data["ATR"]/(c+1e-9)
    data["Vol_10"] =data["Ret_1d"].rolling(10).std()
    data["Vol_20"] =data["Ret_1d"].rolling(20).std()
    data["Vol_ratio"]=data["Vol_10"]/(data["Vol_20"]+1e-9)
    data["Vol_R10"]=v/(v.rolling(10).mean()+1e-9)

    # Candle structure
    o=safe_s(df,"Open"); rng=h-l
    data["Body_pct"]=(c-o).abs()/(rng+1e-9)
    data["Bull_c"]  =(c>o).astype(int)
    data["HH_5"]    =(h>h.shift(1)).rolling(5).sum()
    data["LL_5"]    =(l<l.shift(1)).rolling(5).sum()
    data["Consec_bull"]=(c>c.shift(1)).astype(int).rolling(3).sum()
    data["Consec_bear"]=(c<c.shift(1)).astype(int).rolling(3).sum()

    # PCR proxy
    data["PCR_proxy"]=(v*(c<c.shift()).astype(int)).rolling(5).sum()/\
                      ((v*(c>c.shift()).astype(int)).rolling(5).sum()+1e-9)

    # News sentiment (constant for the day — injected from scraper)
    data["News_score"] = float(news_score) / 100.0

    # Lags
    for lag in [1,2,3,5]:
        data[f"c_lag{lag}"]=c.shift(lag)
        data[f"r_lag{lag}"]=data["Ret_1d"].shift(lag)
        data[f"rsi_lag{lag}"]=data["RSI"].shift(lag)

    # Day of week (Friday = more volatile near close)
    try: data["DayOfWeek"]=pd.to_datetime(data.index).dayofweek
    except: data["DayOfWeek"]=0

    # Expiry week proxy (Bank Nifty expires weekly on Thursday)
    try:
        dts=pd.to_datetime(data.index)
        days_to_thu=((3-dts.dayofweek)%7)
        data["DaysToExpiry"]=days_to_thu.values
    except: data["DaysToExpiry"]=3

    return data


ALL_FEATURES = [
    "Ret_1d","Ret_2d","Ret_3d","Ret_5d","Ret_10d","Ret_20d",
    "MA5_slp","MA20_slp","MA50_slp","P_MA20","P_MA50","P_MA200","MA50_200","EMA9_21",
    "MACD","MACD_sig","MACD_hist","MACD_slp",
    "RSI","RSI_slp","RSI_9","Stoch_K","Stoch_D","Stoch_diff",
    "ADX","DI_diff","PDI","NDI","ST_dir","P_vs_ST",
    "BB_pctb","BB_width","BB_squeeze","CCI","ATR_pct",
    "Vol_10","Vol_20","Vol_ratio","Vol_R10",
    "Body_pct","Bull_c","HH_5","LL_5","Consec_bull","Consec_bear","PCR_proxy",
    "VIX","VIX_change","VIX_vs_avg",
    "SP500_ret","SP500_3d","Nasdaq_ret","Nasdaq_3d",
    "Oil_ret","Oil_3d","Gold_ret","Gold_3d","Nikkei_ret","Nikkei_3d",
    "News_score",
    "DayOfWeek","DaysToExpiry",
    "c_lag1","c_lag2","c_lag3","c_lag5",
    "r_lag1","r_lag2","r_lag3","r_lag5",
    "rsi_lag1","rsi_lag2","rsi_lag3",
]

# ══════════════════════════════════════════════════════════════════════════════
# HMM REGIME DETECTION
# ══════════════════════════════════════════════════════════════════════════════

class RegimeDetector:
    def __init__(self): self.model=None; self.regime_map={}

    def fit(self, data):
        ret=data["Ret_1d"].dropna().values.reshape(-1,1)
        vol=data["Vol_10"].dropna().values.reshape(-1,1)
        n=min(len(ret),len(vol))
        feat=np.hstack([ret[-n:],vol[-n:]])
        self.model=hmm.GaussianHMM(n_components=2,covariance_type="full",
                                    n_iter=300,random_state=42)
        self.model.fit(feat)
        means=self.model.means_[:,1]
        self.regime_map={0:"TRENDING",1:"RANGING"} if means[0]>means[1] \
                   else {0:"RANGING",1:"TRENDING"}
        print(f"   ✅ HMM trained → {self.regime_map}")

    def predict(self, data):
        if self.model is None: return "UNKNOWN"
        try:
            ret=data["Ret_1d"].dropna().values.reshape(-1,1)
            vol=data["Vol_10"].dropna().values.reshape(-1,1)
            n=min(len(ret),len(vol))
            feat=np.hstack([ret[-n:],vol[-n:]])
            return self.regime_map.get(int(self.model.predict(feat)[-1]),"UNKNOWN")
        except: return "UNKNOWN"

regime_det = RegimeDetector()

# ══════════════════════════════════════════════════════════════════════════════
# ML ENSEMBLE  (LightGBM + XGBoost + LR, walk-forward CV)
# ══════════════════════════════════════════════════════════════════════════════

class Ensemble:
    def __init__(self, name):
        self.name=name; self.lgbm=None; self.xgb_m=None
        self.lr=None; self.gb_px=None
        self.scaler=StandardScaler(); self.features=None; self.cv_acc=None

    def train(self, data, features):
        self.features=features
        df=data.copy()
        df["Target"]=(df["Close"].shift(-1)>df["Close"]).astype(int)
        df["Target_Price"]=df["Close"].shift(-1)
        df=df.dropna()
        X=df[features].fillna(0); yd=df["Target"]; yp=df["Target_Price"]
        if len(X)<80: return False

        tscv=TimeSeriesSplit(n_splits=5); accs=[]
        for tr_i,te_i in tscv.split(X):
            m=lgb.LGBMClassifier(n_estimators=200,learning_rate=0.03,num_leaves=31,
                                  min_child_samples=15,subsample=0.8,colsample_bytree=0.8,
                                  random_state=42,verbose=-1)
            m.fit(X.iloc[tr_i],yd.iloc[tr_i],
                  eval_set=[(X.iloc[te_i],yd.iloc[te_i])],
                  callbacks=[lgb.early_stopping(20,verbose=False),
                             lgb.log_evaluation(period=-1)])
            accs.append(accuracy_score(yd.iloc[te_i],m.predict(X.iloc[te_i])))
        self.cv_acc=round(np.mean(accs)*100,1)
        print(f"   [{self.name}] Walk-forward CV: {self.cv_acc}%  {[round(a*100,1) for a in accs]}")

        Xs=self.scaler.fit_transform(X)
        self.lgbm=lgb.LGBMClassifier(n_estimators=500,learning_rate=0.02,num_leaves=31,
                                       min_child_samples=15,subsample=0.8,colsample_bytree=0.8,
                                       reg_alpha=0.1,reg_lambda=0.1,random_state=42,verbose=-1)
        self.lgbm.fit(X,yd)
        self.xgb_m=xgb.XGBClassifier(n_estimators=300,max_depth=6,learning_rate=0.03,
                                       subsample=0.8,colsample_bytree=0.8,
                                       eval_metric="logloss",random_state=42,verbosity=0)
        self.xgb_m.fit(X,yd)
        self.lr=LogisticRegression(C=0.3,max_iter=1000,random_state=42)
        self.lr.fit(Xs,yd)
        self.gb_px=lgb.LGBMRegressor(n_estimators=300,learning_rate=0.02,
                                      num_leaves=31,subsample=0.8,random_state=42,verbose=-1)
        self.gb_px.fit(X,yp)
        return True

    def predict(self, data):
        if self.lgbm is None: return None
        f=data[self.features].iloc[[-1]].fillna(0)
        fs=self.scaler.transform(f)
        lp=self.lgbm.predict_proba(f)[0]
        xp=self.xgb_m.predict_proba(f)[0]
        lrp=self.lr.predict_proba(fs)[0]
        bull=lp[1]*0.5+xp[1]*0.3+lrp[1]*0.2
        bear=lp[0]*0.5+xp[0]*0.3+lrp[0]*0.2
        direction="BUY" if bull>bear else "SELL"
        conf=round(max(bull,bear)*100,1)
        px=float(self.gb_px.predict(f)[0])
        curr=float(data["Close"].iloc[-1])
        return {"direction":direction,"confidence":conf,
                "strength":"STRONG" if conf>=72 else "MODERATE" if conf>=62 else "WEAK",
                "current":curr,"predicted":round(px,2),"cv_accuracy":self.cv_acc}

models={"TRENDING":Ensemble("TRENDING"),"RANGING":Ensemble("RANGING")}
model_all=Ensemble("ALL")

# ══════════════════════════════════════════════════════════════════════════════
# CONSTITUENT BANKS
# ══════════════════════════════════════════════════════════════════════════════

def analyse_banks():
    res=[]; hb=0; hbe=0
    for tkr,(name,wt) in BANKS.items():
        try:
            df=yf.download(tkr,period="30d",auto_adjust=True,progress=False)
            if len(df)<5: continue
            c=safe_s(df,"Close"); h_=safe_s(df,"High"); l_=safe_s(df,"Low")
            curr=float(c.iloc[-1]); prev=float(c.iloc[-2])
            week=float(c.iloc[-6]) if len(c)>=6 else curr
            mon=float(c.iloc[0])
            dc=pct(curr,prev); wc=pct(curr,week); mc=pct(curr,mon)
            r_v=round(float(rsi_f(c).iloc[-1]),1)
            a_v,pd_,nd_=adx_f(h_,l_,c)
            a_v=round(float(a_v.iloc[-1]),1)
            di=round(float((pd_-nd_).iloc[-1]),1)
            ma20=float(c.rolling(20).mean().iloc[-1])
            bull=sum([dc>0,wc>0,mc>0,r_v>50,di>0,curr>ma20])
            trend="🟢 BULL" if bull>=4 else "🔴 BEAR" if bull<=2 else "🟡 MIX"
            is_bull=bull>=4
            if tkr in HEAVYWEIGHTS:
                if is_bull: hb+=1
                else: hbe+=1
            res.append({"name":name,"weight":wt,"price":curr,"day":dc,
                        "week":wc,"month":mc,"rsi":r_v,"adx":a_v,
                        "trend":trend,"is_bull":is_bull,"heavy":tkr in HEAVYWEIGHTS})
        except: pass
    res.sort(key=lambda x:x["weight"],reverse=True)
    nb=sum(1 for b in res if b["is_bull"]); nbe=len(res)-nb
    wbull=sum(b["weight"] for b in res if b["is_bull"])
    wbear=sum(b["weight"] for b in res if not b["is_bull"])
    wb=round((wbull/(wbull+wbear))*100) if (wbull+wbear)>0 else 50
    bias="BULLISH" if wb>=55 else "BEARISH" if wb<=45 else "NEUTRAL"
    return {"banks":res,"n_bull":nb,"n_bear":nbe,"w_breadth":wb,
            "heavy_bull":hb,"heavy_bear":hbe,"sector_bias":bias}

# ══════════════════════════════════════════════════════════════════════════════
# GLOBAL CUES
# ══════════════════════════════════════════════════════════════════════════════

def fetch_global():
    cues={}
    for tkr,nm in GLOBAL_TICKERS.items():
        try:
            df=yf.download(tkr,period="5d",auto_adjust=True,progress=False)
            c=safe_s(df,"Close"); curr=float(c.iloc[-1]); prev=float(c.iloc[-2])
            cues[tkr]={"name":nm,"price":curr,"change":pct(curr,prev)}
        except: cues[tkr]={"name":nm,"price":0,"change":0}
    w={"^GSPC":3,"^IXIC":2,"^DJI":2,"^HSI":1,"^N225":1,"GC=F":0.5,"CL=F":-0.5}
    score=sum(cues.get(t,{}).get("change",0)*wt for t,wt in w.items())
    tw=sum(abs(wt) for wt in w.values())
    comp=round(score/tw,3) if tw>0 else 0
    bias=("🟢 BULLISH GLOBAL" if comp>0.3 else "🟡 MILD BULL" if comp>0.1
          else "🔴 BEARISH GLOBAL" if comp<-0.3 else "🟡 MILD BEAR" if comp<-0.1
          else "⚪ NEUTRAL GLOBAL")
    return {"cues":cues,"composite":comp,"bias":bias}

def fetch_vix():
    try:
        vx=yf.download("^INDIAVIX",period="5d",auto_adjust=True,progress=False)["Close"]
        if isinstance(vx,pd.DataFrame): vx=vx.iloc[:,0]
        curr=float(vx.iloc[-1]); prev=float(vx.iloc[-2]); chg=pct(curr,prev)
        sig=("🟢 STRONG BULL" if chg<-3 else "🟢 MILD BULL" if chg<-1
             else "🔴 STRONG BEAR" if chg>3 else "🔴 MILD BEAR" if chg>1 else "⚪ FLAT")
        return {"level":curr,"change":chg,"signal":sig}
    except: return {"level":15.0,"change":0,"signal":"⚪ UNKNOWN"}

# ══════════════════════════════════════════════════════════════════════════════
# INTRADAY PATTERNS (15-min)
# ══════════════════════════════════════════════════════════════════════════════

def find_patterns(df):
    pats=[]
    if len(df)<4: return pats
    c=safe_s(df,"Close").values; h=safe_s(df,"High").values
    l=safe_s(df,"Low").values;   o=safe_s(df,"Open").values
    v=safe_s(df,"Volume").values; n=len(c)

    if n>=2:
        if c[-2]<o[-2] and c[-1]>o[-1] and o[-1]<=c[-2] and c[-1]>=o[-2]:
            pats.append({"name":"🕯️ Bullish Engulfing","sig":"BUY","conf":80,"desc":"Buyers overwhelmed sellers completely"})
        if c[-2]>o[-2] and c[-1]<o[-1] and o[-1]>=c[-2] and c[-1]<=o[-2]:
            pats.append({"name":"🕯️ Bearish Engulfing","sig":"SELL","conf":80,"desc":"Sellers overwhelmed buyers completely"})
    if n>=1:
        body=abs(c[-1]-o[-1]); lw=min(c[-1],o[-1])-l[-1]; hw=h[-1]-max(c[-1],o[-1])
        if body>0:
            if lw>2*body and hw<body*0.3:
                pats.append({"name":"🔨 Hammer","sig":"BUY","conf":70,"desc":"Buyers rejected lower prices strongly"})
            if hw>2*body and lw<body*0.3:
                pats.append({"name":"🌠 Shooting Star","sig":"SELL","conf":70,"desc":"Sellers rejected higher prices strongly"})
    if n>=3:
        if all(c[-(i+1)]>o[-(i+1)] for i in range(3)):
            pats.append({"name":"📈 3 Consecutive Green","sig":"BUY","conf":65,"desc":"Sustained buying — trend continuation"})
        if all(c[-(i+1)]<o[-(i+1)] for i in range(3)):
            pats.append({"name":"📉 3 Consecutive Red","sig":"SELL","conf":65,"desc":"Sustained selling — trend continuation"})
    if n>=4:
        if h[-1]>h[-2]>h[-3] and l[-1]>l[-2]>l[-3]:
            pats.append({"name":"📈 Higher High-Higher Low","sig":"BUY","conf":75,"desc":"Textbook uptrend structure"})
        if h[-1]<h[-2]<h[-3] and l[-1]<l[-2]<l[-3]:
            pats.append({"name":"📉 Lower High-Lower Low","sig":"SELL","conf":75,"desc":"Textbook downtrend structure"})
    if n>=2:
        if h[-1]<h[-2] and l[-1]>l[-2]:
            pats.append({"name":"📦 Inside Bar","sig":"WAIT","conf":60,"desc":"Consolidation — breakout coming, wait for direction"})
    if n>=5:
        av=np.mean(v[-6:-1])
        if v[-1]>av*2.5:
            sd_="BUY" if c[-1]>o[-1] else "SELL"
            pats.append({"name":f"⚡ Vol Spike ({'Bullish' if sd_=='BUY' else 'Bearish'})","sig":sd_,
                         "conf":85,"desc":f"Volume {v[-1]/av:.1f}x average — institutional activity"})
    if n>=2:
        gap=((o[-1]-c[-2])/c[-2])*100
        if gap>0.3:
            pats.append({"name":"🚀 Gap Up Open","sig":"BUY","conf":70,"desc":f"Opened {gap:.1f}% above prev close"})
        elif gap<-0.3:
            pats.append({"name":"💥 Gap Down Open","sig":"SELL","conf":70,"desc":f"Opened {abs(gap):.1f}% below prev close"})
    return pats

# ══════════════════════════════════════════════════════════════════════════════
# SESSION TRACKER
# ══════════════════════════════════════════════════════════════════════════════

class Session:
    def reset(self):
        self.vwap=None; self.or_high=None; self.or_low=None
        self.or_set=False; self.orb_det=False; self.orb_dir=None
        self.last_date=None; self.morning_done=False
        self.trades_today=0; self.force_exit_alerted=False
    def __init__(self): self.reset()
    def set_or(self,h,l):
        if not self.or_set:
            self.or_high=h; self.or_low=l; self.or_set=True
            day_log.or_high=h; day_log.or_low=l; day_log.or_range=round(h-l,2)
    def check_orb(self,p):
        if not self.or_set or self.orb_det: return
        span=self.or_high-self.or_low
        if p>self.or_high+span*0.003: self.orb_det=True; self.orb_dir="BULLISH"
        elif p<self.or_low-span*0.003: self.orb_det=True; self.orb_dir="BEARISH"

session=Session()

# ══════════════════════════════════════════════════════════════════════════════
# CONFLUENCE ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def calc_confluence(ml_sig, sess, banks, patterns, gc, vix_i, oc, regime, news):
    bp=0; bep=0; checks=[]
    def chk(n,s,d): checks.append((n,s,d))

    # 1. ML Ensemble — weight 3
    if ml_sig["confidence"]>=MIN_CONFIDENCE:
        if ml_sig["direction"]=="BUY":  bp+=3; chk("ML Ensemble","🟢 BULL",f'{ml_sig["confidence"]}% conf | CV={ml_sig["cv_accuracy"]}%')
        else:                            bep+=3; chk("ML Ensemble","🔴 BEAR",f'{ml_sig["confidence"]}% conf | CV={ml_sig["cv_accuracy"]}%')
    else: chk("ML Ensemble","⚪ WEAK",f'{ml_sig["confidence"]}% — below threshold')

    # 2. ADX trend filter — weight 2
    a=ml_sig.get("adx",0); di=ml_sig.get("di_diff",0)
    if a>=MIN_ADX:
        if di>0: bp+=2; chk("ADX/DI","🟢 BULL",f'ADX={a} +DI leads → trending up')
        else:    bep+=2; chk("ADX/DI","🔴 BEAR",f'ADX={a} -DI leads → trending down')
    else: chk("ADX/DI","⚪ RANGING",f'ADX={a} — choppy market, caution')

    # 3. Supertrend — weight 2
    std=ml_sig.get("st_dir",0)
    if std==-1: bp+=2; chk("Supertrend","🟢 BULL","Price above Supertrend line")
    elif std==1: bep+=2; chk("Supertrend","🔴 BEAR","Price below Supertrend line")

    # 4. Real PCR from NSE — weight 3
    if oc["success"]:
        pcr=oc["pcr"]
        if pcr>=1.5:    bp+=3; chk("Real PCR","🟢 STRONG BULL",f'PCR={pcr} — institutions hedging (very bullish)')
        elif pcr>=1.2:  bp+=2; chk("Real PCR","🟢 BULL",f'PCR={pcr} — more puts than calls')
        elif pcr>=0.9:  bp+=1; chk("Real PCR","🟢 MILD BULL",f'PCR={pcr}')
        elif pcr>=0.7:  bep+=1; chk("Real PCR","🔴 MILD BEAR",f'PCR={pcr}')
        elif pcr>=0.5:  bep+=2; chk("Real PCR","🔴 BEAR",f'PCR={pcr} — call-heavy (over-optimism)')
        else:           bep+=3; chk("Real PCR","🔴 STRONG BEAR",f'PCR={pcr} — extreme call buying')
    else: chk("Real PCR","⚪ N/A","NSE options data unavailable")

    # 5. OI trend (change in OI) — weight 2
    if oc["success"]:
        oit=oc["oi_trend"]
        if oit=="BULLISH":   bp+=2; chk("OI Trend","🟢 BULL","PE building + CE unwinding — bulls loading")
        elif oit=="BEARISH": bep+=2; chk("OI Trend","🔴 BEAR","CE building + PE unwinding — bears loading")
        else:                chk("OI Trend","⚪ NEUTRAL",oit)

    # 6. Max Pain pull — weight 1
    if oc["success"] and oc["max_pain"]>0 and ml_sig.get("current",0)>0:
        mp=oc["max_pain"]; spot=ml_sig["current"]
        dist=((mp-spot)/spot)*100
        if dist>0.5:    bp+=1; chk("Max Pain","🟢 BULL",f'Max pain ₹{mp:,} is {dist:+.1f}% above — price attracted up')
        elif dist<-0.5: bep+=1; chk("Max Pain","🔴 BEAR",f'Max pain ₹{mp:,} is {dist:+.1f}% below — price attracted down')
        else:           chk("Max Pain","⚪ AT PAIN",f'Spot ≈ max pain ₹{mp:,} — limited move expected')

    # 7. News Sentiment — weight 2
    ns=news["score"]
    if ns>=30:    bp+=2; chk("News Sentiment","🟢 BULL",f'Score={ns} — {news["bull_count"]} bullish headlines')
    elif ns>=10:  bp+=1; chk("News Sentiment","🟢 MILD",f'Score={ns}')
    elif ns<=-30: bep+=2; chk("News Sentiment","🔴 BEAR",f'Score={ns} — {news["bear_count"]} bearish headlines')
    elif ns<=-10: bep+=1; chk("News Sentiment","🔴 MILD",f'Score={ns}')
    else:         chk("News Sentiment","⚪ NEUTRAL",f'Score={ns} — balanced news')

    # 8. Global cues — weight 2
    gc_c=gc["composite"]
    if gc_c>0.3:    bp+=2; chk("Global Markets","🟢 BULL",f'Score={gc_c:+.3f} — US/Asia up')
    elif gc_c>0.1:  bp+=1; chk("Global Markets","🟢 MILD",f'Score={gc_c:+.3f}')
    elif gc_c<-0.3: bep+=2; chk("Global Markets","🔴 BEAR",f'Score={gc_c:+.3f} — US/Asia down')
    elif gc_c<-0.1: bep+=1; chk("Global Markets","🔴 MILD",f'Score={gc_c:+.3f}')
    else:           chk("Global Markets","⚪ FLAT",f'Score={gc_c:+.3f}')

    # 9. VIX change direction — weight 2
    vc=vix_i["change"]
    if vc<-3:    bp+=2; chk("VIX Direction","🟢 BULL",f'{vc:+.1f}% — fear dropping sharply')
    elif vc<-1:  bp+=1; chk("VIX Direction","🟢 MILD",f'{vc:+.1f}% — fear easing')
    elif vc>3:   bep+=2; chk("VIX Direction","🔴 BEAR",f'{vc:+.1f}% — fear rising sharply')
    elif vc>1:   bep+=1; chk("VIX Direction","🔴 MILD",f'{vc:+.1f}% — fear increasing')
    else:        chk("VIX Direction","⚪ FLAT",f'{vc:+.1f}%')

    # 10. Bank breadth — weight 2
    wb=banks["w_breadth"]; hb_=banks["heavy_bull"]; hbe_=banks["heavy_bear"]
    if wb>=65:   bp+=2; chk("Bank Breadth","🟢 BULL",f'{wb}% weighted bull — strong sector')
    elif wb>=55: bp+=1; chk("Bank Breadth","🟢 MILD",f'{wb}%')
    elif wb<=35: bep+=2; chk("Bank Breadth","🔴 BEAR",f'{wb}% — weak sector')
    elif wb<=45: bep+=1; chk("Bank Breadth","🔴 MILD",f'{wb}%')
    else:        chk("Bank Breadth","⚪ MIX",f'{wb}% — balanced')

    # 11. Heavyweights — weight 2
    if hb_==3:     bp+=2; chk("Heavyweights","🟢 ALL BULL","HDFC+ICICI+Kotak all bullish → index must follow")
    elif hb_==2:   bp+=1; chk("Heavyweights","🟢 2/3 BULL","2 of 3 top banks bullish")
    elif hbe_==3:  bep+=2; chk("Heavyweights","🔴 ALL BEAR","HDFC+ICICI+Kotak all bearish → index must follow")
    elif hbe_==2:  bep+=1; chk("Heavyweights","🔴 2/3 BEAR","2 of 3 top banks bearish")
    else:          chk("Heavyweights","⚪ SPLIT","Mixed signals — index direction unclear")

    # 12. 15-min patterns — weight 2
    pb=[p for p in patterns if p["sig"]=="BUY"  and p["conf"]>=70]
    pe_=[p for p in patterns if p["sig"]=="SELL" and p["conf"]>=70]
    if len(pb)>=2:    bp+=2; chk("15m Patterns","🟢 BULL",f'{len(pb)} bullish patterns confirmed')
    elif len(pb)==1:  bp+=1; chk("15m Patterns","🟢 BULL",pb[0]["name"])
    elif len(pe_)>=2: bep+=2; chk("15m Patterns","🔴 BEAR",f'{len(pe_)} bearish patterns confirmed')
    elif len(pe_)==1: bep+=1; chk("15m Patterns","🔴 BEAR",pe_[0]["name"])
    else:             chk("15m Patterns","⚪ NONE","No confirmed intraday patterns")

    # 13. Classic indicators
    r=ml_sig.get("rsi",50)
    if r<40:    bp+=1; chk("RSI","🟢 BULL",f'{r} — oversold bounce expected')
    elif r>60:  bep+=1; chk("RSI","🔴 BEAR",f'{r} — overbought pullback risk')
    else:       chk("RSI","⚪ NEUTRAL",f'{r}')

    sk=ml_sig.get("stoch_k",50)
    if sk<25:   bp+=1; chk("Stoch RSI","🟢 BULL",f'{sk} — deeply oversold')
    elif sk>75: bep+=1; chk("Stoch RSI","🔴 BEAR",f'{sk} — deeply overbought')
    else:       chk("Stoch RSI","⚪ MID",f'{sk}')

    mh=ml_sig.get("macd_hist",0)
    if mh>0: bp+=1; chk("MACD Hist","🟢 BULL",f'Histogram {mh:.1f} positive')
    else:    bep+=1; chk("MACD Hist","🔴 BEAR",f'Histogram {mh:.1f} negative')

    bb=ml_sig.get("bb_pctb",0.5)
    if bb<0.2:  bp+=1; chk("Bollinger %B","🟢 BULL",f'{bb:.2f} near lower band')
    elif bb>0.8: bep+=1; chk("Bollinger %B","🔴 BEAR",f'{bb:.2f} near upper band')
    else:        chk("Bollinger %B","⚪ MID",f'{bb:.2f}')

    # VWAP
    if sess.vwap and ml_sig.get("current",0)>0:
        vd=((ml_sig["current"]-sess.vwap)/sess.vwap)*100
        if vd>0.15:    bp+=1; chk("VWAP","🟢 BULL",f'{vd:+.2f}% above VWAP')
        elif vd<-0.15: bep+=1; chk("VWAP","🔴 BEAR",f'{vd:+.2f}% below VWAP')
        else:          chk("VWAP","⚪ AT",f'Within ±0.15% of VWAP')

    # ORB
    if sess.orb_det:
        if sess.orb_dir=="BULLISH": bp+=1; chk("ORB Breakout","🟢 BULL","Confirmed breakout above opening range")
        else:                        bep+=1; chk("ORB Breakout","🔴 BEAR","Confirmed breakdown below opening range")

    total=bp+bep
    if total==0: return {"score":50,"direction":"NEUTRAL","bull_pts":0,"bear_pts":0,"checks":checks,"tradeable":False}
    if bp>bep:   score=round((bp/total)*100); direction="BUY"
    else:        score=round((bep/total)*100); direction="SELL"

    tradeable = (score>=MIN_CONFLUENCE and
                 ml_sig.get("adx",0)>=MIN_ADX and
                 ml_sig.get("confidence",0)>=MIN_CONFIDENCE and
                 session.trades_today < MAX_TRADES_DAY)

    return {"score":score,"direction":direction,"bull_pts":bp,
            "bear_pts":bep,"checks":checks,"tradeable":tradeable}

# ══════════════════════════════════════════════════════════════════════════════
# MORNING BRIEFING  (8:30 AM — before market opens)
# ══════════════════════════════════════════════════════════════════════════════


# ══════════════════════════════════════════════════════════════════════════════
# ALERT / NOTIFICATION ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def send_telegram(message):
    """Send message to your Telegram phone via bot. Free & instant."""
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        return False
    try:
        url  = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
        data = {"chat_id": TELEGRAM_CHAT_ID, "text": message,
                "parse_mode": "HTML"}
        resp = requests.post(url, data=data, timeout=10)
        if resp.status_code == 200:
            print("  📱  Telegram alert sent!")
            return True
        else:
            print(f"  ⚠️  Telegram error: {resp.status_code} {resp.text[:100]}")
            return False
    except Exception as e:
        print(f"  ⚠️  Telegram failed: {e}")
        return False


def send_sms(message):
    """Send actual SMS via Fast2SMS (free tier available)."""
    if not FAST2SMS_API_KEY or not YOUR_MOBILE:
        return False
    try:
        # Fast2SMS supports up to 160 chars per SMS — truncate if needed
        sms_text = message[:160] if len(message) > 160 else message
        url = "https://www.fast2sms.com/dev/bulkV2"
        headers = {"authorization": FAST2SMS_API_KEY}
        params  = {
            "variables_values": sms_text,
            "route"           : "q",        # quick transactional route
            "numbers"         : YOUR_MOBILE,
        }
        resp = requests.get(url, headers=headers, params=params, timeout=10)
        data = resp.json()
        if data.get("return"):
            print("  📲  SMS sent!")
            return True
        else:
            print(f"  ⚠️  SMS error: {data.get('message','unknown')}")
            return False
    except Exception as e:
        print(f"  ⚠️  SMS failed: {e}")
        return False


def send_alert(message, sms_message=None):
    """     Send alert to phone via Telegram AND/OR SMS.     sms_message = shorter version for SMS (optional).     """
    tg_sent  = send_telegram(message)
    sms_text = sms_message if sms_message else message[:160]
    sms_sent = send_sms(sms_text)
    if not tg_sent and not sms_sent:
        print("  ℹ️  No alert sent (configure TELEGRAM_TOKEN or FAST2SMS_API_KEY)")


def alert_trade_signal(direction, strike, typ, entry, sl, t1, t2, t3,
                        score, spot, regime, lots, rr, expiry, mode):
    """Formats and sends trade signal alert to phone."""
    if not ALERT_ON_SIGNAL:
        return
    now_str = datetime.now(IST).strftime("%H:%M")
    emoji   = "🟢" if direction == "BUY" else "🔴"
    action  = "BUY CALL (CE)" if direction == "BUY" else "BUY PUT (PE)"
    mode_tag= "[PAPER]" if mode else "[LIVE]"

    # Full Telegram message (supports emojis, HTML)
    tg_msg = (
        f"<b>{emoji} BANKNIFTY SIGNAL {mode_tag}</b> — {now_str}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"<b>ACTION   :</b> {action}\n"
        f"<b>CONTRACT :</b> BANKNIFTY {strike} {typ}\n"
        f"<b>EXPIRY   :</b> {expiry}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"<b>ENTRY    :</b> Rs.{entry:.0f}\n"
        f"<b>LOTS     :</b> {lots} lot ({lots*25} qty)\n"
        f"<b>INVEST   :</b> Rs.{round(entry*lots*25):,}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"<b>TARGET 1 :</b> Rs.{t1:.0f}  (exit 30%)\n"
        f"<b>TARGET 2 :</b> Rs.{t2:.0f}  (exit 40%)\n"
        f"<b>TARGET 3 :</b> Rs.{t3:.0f}  (exit 30%)\n"
        f"<b>STOP LOSS:</b> Rs.{sl:.0f}  SET IMMEDIATELY\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"<b>R:R      :</b> 1:{rr:.1f}\n"
        f"<b>SCORE    :</b> {score}/100\n"
        f"<b>REGIME   :</b> {regime}\n"
        f"<b>SPOT     :</b> Rs.{spot:,.0f}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"After T1 hit: Move SL to entry (risk free)\n"
        f"EXIT ALL by 3:15 PM — no exceptions"
    )

    # Short SMS version (160 chars)
    sms_msg = (
        f"{emoji}BNIFTY {direction} {strike}{typ} "
        f"Entry:{entry:.0f} SL:{sl:.0f} "
        f"T1:{t1:.0f} T2:{t2:.0f} "
        f"Score:{score} RR:1:{rr:.1f} {mode_tag}"
    )

    send_alert(tg_msg, sms_msg)


def alert_morning_briefing(day_bias, global_bias, vix_level, vix_chg,
                            news_bias, bank_breadth, events):
    """Send morning summary to phone."""
    if not ALERT_MORNING_BRIEF:
        return
    now_str  = datetime.now(IST).strftime("%d %b %Y")
    evt_text = ""
    if events:
        evt_text = "\n⚠️ EVENT: " + events[0][1][:60]

    tg_msg = (
        f"<b>☀️ BANKNIFTY MORNING BRIEF</b> — {now_str}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"<b>DAY BIAS    :</b> {day_bias}\n"
        f"<b>GLOBAL      :</b> {global_bias}\n"
        f"<b>INDIA VIX   :</b> {vix_level} ({vix_chg:+.1f}%)\n"
        f"<b>NEWS        :</b> {news_bias}\n"
        f"<b>BANK SECTOR :</b> {bank_breadth}% bullish\n"
        f"{evt_text}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"Market opens 9:15 AM\n"
        f"DO NOT trade before 9:45 AM\n"
        f"Max trades today: {MAX_TRADES_DAY}"
    )

    sms_msg = (
        f"BN Morning: {day_bias[:20]} | "
        f"VIX:{vix_level}({vix_chg:+.1f}%) | "
        f"News:{news_bias[:15]} | "
        f"Banks:{bank_breadth}%bull"
    )

    send_alert(tg_msg, sms_msg)


def alert_target_hit(target_num, spot, premium, profit, contract):
    """Alert when a target level is approached."""
    if not ALERT_ON_TARGET:
        return
    now_str = datetime.now(IST).strftime("%H:%M")
    tg_msg  = (
        f"<b>🎯 TARGET {target_num} HIT!</b> — {now_str}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"<b>Contract :</b> {contract}\n"
        f"<b>Spot Now :</b> Rs.{spot:,.0f}\n"
        f"<b>Premium  :</b> Rs.{premium:.0f}\n"
        f"<b>Profit   :</b> Rs.{profit:,}/lot\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        + ("EXIT 30% of position NOW" if target_num == 1
           else "EXIT 40% of position NOW" if target_num == 2
           else "EXIT remaining 30% — FULL EXIT")
        + "\nMove SL to entry after T1"
    )
    sms_msg = f"T{target_num} HIT! {contract} Spot:{spot:.0f} Profit:Rs.{profit}/lot"
    send_alert(tg_msg, sms_msg)


def alert_force_exit(spot):
    """3:15 PM mandatory exit reminder."""
    if not ALERT_FORCE_EXIT:
        return
    tg_msg = (
        f"<b>🔔 3:15 PM — MANDATORY EXIT</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"CLOSE ALL open BANKNIFTY positions NOW\n"
        f"Spot: Rs.{spot:,.0f}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"Groww: F&O > Positions > Close Position\n"
        f"No exceptions. Even if in loss."
    )
    sms_msg = f"3:15PM EXIT ALERT: Close ALL BankNifty positions NOW. Spot:{spot:.0f}"
    send_alert(tg_msg, sms_msg)


def alert_sl_warning(spot, sl_level, contract):
    """Alert when price is approaching stop loss."""
    if not ALERT_ON_SL:
        return
    now_str = datetime.now(IST).strftime("%H:%M")
    tg_msg  = (
        f"<b>⚠️ SL WARNING</b> — {now_str}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"<b>Contract :</b> {contract}\n"
        f"<b>Spot Now :</b> Rs.{spot:,.0f}\n"
        f"<b>Your SL  :</b> Rs.{sl_level:,.0f}\n"
        f"Price is near your stop loss!\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"If SL is triggered: EXIT immediately\n"
        f"Do NOT hold hoping it will recover"
    )
    sms_msg = f"SL WARNING: {contract} Spot:{spot:.0f} near SL:{sl_level:.0f} - Stay alert!"
    send_alert(tg_msg, sms_msg)


def test_alerts():
    """Test both Telegram and SMS on startup."""
    print("\n  📱  Testing alert connections…")
    test_msg = (
        "<b>✅ BankNifty Alert System CONNECTED</b>\n"
        "You will now receive trade signals on this phone.\n"
        "Paper trade mode is ON — no real money at risk.\n"
        "Good luck! 🎯"
    )
    sms_test = "BankNifty Alerts connected! You will get trade signals here."
    tg  = send_telegram(test_msg)
    sms = send_sms(sms_test)
    if not tg and not sms:
        print("  ℹ️  No alert channels configured.")
        print("      Fill TELEGRAM_TOKEN or FAST2SMS_API_KEY above to enable.")
    else:
        print("  ✅  Alert test complete!")



# ══════════════════════════════════════════════════════════════════════════════
# EXCEL TRADE TRACKER  — auto-logs everything, learns from results
# ══════════════════════════════════════════════════════════════════════════════

import subprocess as _sp
_sp.check_call([sys.executable,"-m","pip","install","-q","openpyxl",
                "--break-system-packages"])

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# EXCEL_PATH defined in server header above

def _thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s,right=s,top=s,bottom=s)

def _style(cell, bg="FFFFFF", bold=False, color="000000",
           align="center", size=9):
    cell.font      = Font(name="Arial", bold=bold, size=size, color=color)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=True)
    cell.border    = _thin_border()

def _init_excel():
    """Create fresh Excel workbook with all sheets."""
    wb = Workbook()

    # ── TRADE_JOURNAL ─────────────────────────────────────────────────────
    ws = wb.active; ws.title = "TRADE_JOURNAL"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:T1")
    c = ws["A1"]
    c.value = "BANK NIFTY TRADE JOURNAL — Auto-filled by v8 System"
    _style(c, bg="1A5276", bold=True, color="FFFFFF", size=12)
    ws.row_dimensions[1].height = 28

    hdrs = ["Date","Time","Dir","Strike","Type","Entry Prem","SL Prem",
            "T1 Prem","T2 Prem","T3 Prem","Spot Entry","SL Spot",
            "Lots","Invest (Rs)","Outcome","Exit Prem","P&L (Rs)",
            "Score","Regime","Notes"]
    bg_h = "154360"
    for i,h in enumerate(hdrs,1):
        c = ws.cell(row=2,column=i); c.value=h
        _style(c,bg=bg_h,bold=True,color="FFFFFF",size=9)
        ws.column_dimensions[get_column_letter(i)].width = [
            12,8,8,8,6,9,9,9,9,9,11,10,6,12,10,9,12,7,10,18][i-1]
    ws.freeze_panes = "C3"

    # ── PERFORMANCE ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("PERFORMANCE")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:C1")
    c = ws2["A1"]
    c.value = "PERFORMANCE — Live stats from Trade Journal"
    _style(c,bg="154360",bold=True,color="FFFFFF",size=12)
    ws2.row_dimensions[1].height=28

    stats = [
        ("Total Trades",    "=COUNTA(TRADE_JOURNAL!A3:A99999)-1"),
        ("Wins",            '=COUNTIF(TRADE_JOURNAL!Q3:Q99999,">"&0)'),
        ("Losses",          '=COUNTIF(TRADE_JOURNAL!Q3:Q99999,"<"&0)'),
        ("Win Rate %",      "=IFERROR(B4/B3*100,0)"),
        ("Total P&L (Rs)",  "=SUM(TRADE_JOURNAL!Q3:Q99999)"),
        ("Avg Win (Rs)",    '=IFERROR(AVERAGEIF(TRADE_JOURNAL!Q3:Q99999,">"&0,TRADE_JOURNAL!Q3:Q99999),0)'),
        ("Avg Loss (Rs)",   '=IFERROR(AVERAGEIF(TRADE_JOURNAL!Q3:Q99999,"<"&0,TRADE_JOURNAL!Q3:Q99999),0)'),
        ("Best Trade (Rs)", "=IFERROR(MAX(TRADE_JOURNAL!Q3:Q99999),0)"),
        ("Worst Trade (Rs)","=IFERROR(MIN(TRADE_JOURNAL!Q3:Q99999),0)"),
        ("Avg R:R",         "=IFERROR(ABS(B8/B9),0)"),
        ("SL Hit count",    '=COUNTIF(TRADE_JOURNAL!O3:O99999,"SL HIT")'),
        ("T1 Hit count",    '=COUNTIF(TRADE_JOURNAL!O3:O99999,"T1 HIT")'),
        ("T2 Hit count",    '=COUNTIF(TRADE_JOURNAL!O3:O99999,"T2 HIT")'),
        ("T3 Hit count",    '=COUNTIF(TRADE_JOURNAL!O3:O99999,"T3 HIT")'),
    ]
    bgs = ["D6EAF8","D5F5E3","FADBD8","D6EAF8","D6EAF8",
           "D5F5E3","FADBD8","D5F5E3","FADBD8","D6EAF8",
           "FADBD8","D5F5E3","D5F5E3","D5F5E3"]
    for i,(lbl,fml) in enumerate(stats):
        r=i+3
        lc=ws2.cell(row=r,column=1); vc=ws2.cell(row=r,column=2)
        lc.value=lbl; vc.value=fml
        _style(lc,bg=bgs[i],align="left",size=10)
        _style(vc,bg=bgs[i],bold=True,size=10)
        ws2.row_dimensions[r].height=20
    ws2["B6"].number_format="0.0"
    for r in [7,8,9,10,11]:
        ws2.cell(row=r,column=2).number_format="#,##0"
    ws2.column_dimensions["A"].width=22
    ws2.column_dimensions["B"].width=16

    # ── PRICE_LOG ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("PRICE_LOG")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:G1")
    c=ws3["A1"]; c.value="PRICE LOG — Spot price every 2 min during open trade"
    _style(c,bg="1A5276",bold=True,color="FFFFFF",size=11)
    ws3.row_dimensions[1].height=26
    for i,h in enumerate(["Date","Time","Spot","Entry Spot","T1 Spot","SL Spot","Dist to SL %"],1):
        c=ws3.cell(row=2,column=i); c.value=h
        _style(c,bg="1A5276",bold=True,color="FFFFFF",size=9)
    widths2=[12,8,12,12,12,12,10]
    for i,w in enumerate(widths2,1):
        ws3.column_dimensions[get_column_letter(i)].width=w
    ws3.freeze_panes="C3"

    # ── ML_FEEDBACK ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("ML_FEEDBACK")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:M1")
    c=ws4["A1"]; c.value="ML FEEDBACK — Features + outcome for model retraining"
    _style(c,bg="4A235A",bold=True,color="FFFFFF",size=11)
    ws4.row_dimensions[1].height=26
    ml_h=["Date","Time","RSI","ADX","Stoch_K","MACD_H",
          "BB_pctb","VIX","VIX_chg","SP500","News","Score","Outcome"]
    for i,h in enumerate(ml_h,1):
        c=ws4.cell(row=2,column=i); c.value=h
        _style(c,bg="4A235A",bold=True,color="FFFFFF",size=9)
        ws4.column_dimensions[get_column_letter(i)].width=10
    ws4.freeze_panes="C3"

    # ── DAILY_SUMMARY ─────────────────────────────────────────────────────
    ws5 = wb.create_sheet("DAILY_SUMMARY")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:I1")
    c=ws5["A1"]; c.value="DAILY SUMMARY — One row per day"
    _style(c,bg="BA4A00",bold=True,color="FFFFFF",size=11)
    ws5.row_dimensions[1].height=26
    for i,h in enumerate(["Date","Signals","Trades","Wins","Losses",
                           "Day P&L","Running Total","Bias","Notes"],1):
        c=ws5.cell(row=2,column=i); c.value=h
        _style(c,bg="BA4A00",bold=True,color="FFFFFF",size=9)
    for i,w in enumerate([12,9,9,7,7,14,14,12,25],1):
        ws5.column_dimensions[get_column_letter(i)].width=w
    ws5.freeze_panes="B3"

    wb.save(EXCEL_PATH)
    print(f"   ✅ Excel tracker created: {EXCEL_PATH}")
    return wb


def _load_wb():
    """Load workbook safely, create if missing."""
    import os
    if not os.path.exists(EXCEL_PATH):
        return _init_excel(), True
    try:
        return load_workbook(EXCEL_PATH), False
    except:
        return _init_excel(), True


def excel_log_signal(date_str, time_str, direction, strike, typ,
                     entry, sl_prem, t1, t2, t3,
                     spot, sl_spot, lots,
                     score, regime,
                     rsi, adx, stoch_k, macd_h, bb_pctb,
                     vix, vix_chg, sp500, news_score):
    """Write a new trade signal row to TRADE_JOURNAL and ML_FEEDBACK."""
    try:
        wb, _ = _load_wb()

        # ── TRADE_JOURNAL ─────────────────────────────────────────────────
        ws  = wb["TRADE_JOURNAL"]
        row = ws.max_row + 1
        if row < 3: row = 3

        invest = round(entry * lots * LOT_SIZE)
        vals = [date_str, time_str, direction, strike, typ,
                round(entry,2), round(sl_prem,2),
                round(t1,2), round(t2,2), round(t3,2),
                round(spot,2), round(sl_spot,2),
                lots, invest,
                "OPEN", "", "",   # Outcome / Exit / P&L filled later
                score, regime, ""]

        bg = "D5F5E3" if direction=="BUY" else "FADBD8"
        for i,v in enumerate(vals,1):
            c = ws.cell(row=row, column=i)
            c.value = v
            _style(c, bg=bg, size=9)

        # Colour outcome column blue (OPEN)
        _style(ws.cell(row=row,column=15), bg="D6EAF8", size=9)

        # Store row number for later outcome update
        # We use Notes column to store row reference
        ws.cell(row=row, column=20).value = f"ROW:{row}"

        # ── ML_FEEDBACK ───────────────────────────────────────────────────
        ws4  = wb["ML_FEEDBACK"]
        row4 = ws4.max_row + 1
        if row4 < 3: row4 = 3

        ml_vals = [date_str, time_str,
                   round(rsi,1), round(adx,1), round(stoch_k,1),
                   round(macd_h,4), round(bb_pctb,3),
                   round(vix,1), round(vix_chg,2),
                   round(sp500,3), round(news_score,0),
                   score, "OPEN"]
        for i,v in enumerate(ml_vals,1):
            c = ws4.cell(row=row4, column=i)
            c.value = v
            _style(c, bg="EDE7F6", size=9)

        wb.save(EXCEL_PATH)
        print(f"   📊 Excel: signal logged (row {row})")
        return row   # return row so we can update outcome later

    except Exception as e:
        print(f"   ⚠️  Excel log error: {e}")
        return None


def excel_update_outcome(signal_row, outcome, exit_prem, pnl):
    """Update outcome when T1/T2/T3 or SL is hit."""
    try:
        wb, _ = _load_wb()
        ws = wb["TRADE_JOURNAL"]
        if signal_row and signal_row <= ws.max_row:
            ws.cell(row=signal_row, column=15).value = outcome   # Outcome
            ws.cell(row=signal_row, column=16).value = round(exit_prem,2)  # Exit prem
            ws.cell(row=signal_row, column=17).value = round(pnl,0)        # P&L
            # Colour row green or red
            bg = "D5F5E3" if pnl>0 else "FADBD8" if pnl<0 else "D6EAF8"
            for col in range(1,21):
                _style(ws.cell(row=signal_row, column=col), bg=bg, size=9)

        # Update ML_FEEDBACK outcome
        ws4 = wb["ML_FEEDBACK"]
        for row in range(3, ws4.max_row+1):
            if ws4.cell(row=row, column=13).value == "OPEN":
                ws4.cell(row=row, column=13).value = outcome
                break

        wb.save(EXCEL_PATH)
        print(f"   📊 Excel: outcome updated → {outcome} | P&L: Rs.{pnl:,.0f}")
    except Exception as e:
        print(f"   ⚠️  Excel outcome error: {e}")


def excel_log_price(date_str, time_str, spot, entry_spot, t1_spot, sl_spot):
    """Log every 2-min price during active trade."""
    try:
        wb, _ = _load_wb()
        ws3  = wb["PRICE_LOG"]
        row  = ws3.max_row + 1
        if row < 3: row = 3
        dist = abs(spot-sl_spot)/sl_spot*100 if sl_spot>0 else 0
        bg   = "FADBD8" if dist<0.5 else "D6EAF8"
        for i,v in enumerate([date_str, time_str,
                               round(spot,2), round(entry_spot,2),
                               round(t1_spot,2), round(sl_spot,2),
                               round(dist,2)], 1):
            c = ws3.cell(row=row, column=i); c.value=v
            _style(c, bg=bg, size=9)
        wb.save(EXCEL_PATH)
    except Exception as e:
        print(f"   ⚠️  Price log error: {e}")


def excel_log_daily(date_str, signals, trades, wins, losses, pnl, bias):
    """Log end-of-day summary."""
    try:
        wb, _ = _load_wb()
        ws5  = wb["DAILY_SUMMARY"]
        row  = ws5.max_row + 1
        if row < 3: row = 3

        # Running total formula
        running = f"=F{row}+G{row-1}" if row > 3 else f"=F{row}"

        vals = [date_str, signals, trades, wins, losses,
                round(pnl,0), running, bias, ""]
        bg = "D5F5E3" if pnl>=0 else "FADBD8"
        for i,v in enumerate(vals,1):
            c = ws5.cell(row=row, column=i); c.value=v
            _style(c, bg=bg, size=9)
        ws5.cell(row=row,column=7).number_format="#,##0"
        ws5.cell(row=row,column=6).number_format="#,##0"

        wb.save(EXCEL_PATH)
        print(f"   📊 Excel: daily summary logged — P&L Rs.{pnl:,.0f}")
    except Exception as e:
        print(f"   ⚠️  Daily log error: {e}")


def excel_retrain_from_feedback():
    """     Read ML_FEEDBACK sheet. If >= 30 completed trades exist,     use them to retrain the ML model with real-world outcomes.     Returns retraining accuracy or None.     """
    try:
        import os
        if not os.path.exists(EXCEL_PATH):
            return None
        wb, _ = _load_wb()
        ws4 = wb["ML_FEEDBACK"]
        rows = []
        for row in range(3, ws4.max_row+1):
            outcome = ws4.cell(row=row, column=13).value
            if outcome and outcome != "OPEN":
                r = {}
                cols = ["Date","Time","RSI","ADX","Stoch_K","MACD_H",
                        "BB_pctb","VIX","VIX_chg","SP500","News","Score","Outcome"]
                for i,col in enumerate(cols,1):
                    r[col] = ws4.cell(row=row,column=i).value
                rows.append(r)

        if len(rows) < 30:
            print(f"   ℹ️  Excel feedback: {len(rows)} trades logged "
                  f"(need 30 to retrain)")
            return None

        print(f"   🔄 Retraining from {len(rows)} real trades in Excel…")
        import pandas as pd
        df = pd.DataFrame(rows)

        # Map outcome to binary: win = T1/T2/T3, loss = SL HIT
        df["target"] = df["Outcome"].apply(
            lambda x: 1 if x in ["T1 HIT","T2 HIT","T3 HIT"] else 0)

        features = ["RSI","ADX","Stoch_K","MACD_H",
                    "BB_pctb","VIX","VIX_chg","SP500","News","Score"]
        df = df.dropna(subset=features+["target"])
        X  = df[features].astype(float)
        y  = df["target"]

        if y.nunique() < 2:
            print("   ℹ️  Not enough outcome variety yet to retrain")
            return None

        from sklearn.model_selection import cross_val_score
        import lightgbm as lgb_
        m = lgb_.LGBMClassifier(n_estimators=100,random_state=42,verbose=-1)
        scores = cross_val_score(m,X,y,cv=min(5,len(X)//5),scoring="accuracy")
        acc = round(scores.mean()*100,1)
        m.fit(X,y)

        print(f"   ✅ Retrained on real trades! Accuracy: {acc}%")
        print(f"   📈 Model now knows YOUR market patterns")

        # Send Telegram alert about retraining
        msg = (f"<b>🧠 ML Model Retrained!</b>\n"
               f"Based on {len(rows)} real trades in your Excel\n"
               f"New accuracy: {acc}%\n"
               f"Model is learning YOUR patterns")
        send_telegram(msg)

        return acc, m

    except Exception as e:
        print(f"   ⚠️  Retrain error: {e}")
        return None


def _init_excel_tracker():
    """Called on startup — creates Excel if not exists."""
    import os
    if not os.path.exists(EXCEL_PATH):
        print("\n  📊  Creating Excel trade tracker…")
        _init_excel()
        print(f"  📊  Excel ready: {EXCEL_PATH}")
        print(f"  📊  Download anytime: Colab → Files (folder icon) → "
              f"banknifty_trades.xlsx")
    else:
        print(f"  📊  Excel tracker found: {EXCEL_PATH}")
        # Check how many trades already logged
        try:
            wb = load_workbook(EXCEL_PATH,data_only=True)
            ws = wb["TRADE_JOURNAL"]
            count = max(0, ws.max_row - 2)
            print(f"  📊  {count} trades already in journal")
            if count >= 30:
                print(f"  🔄  Checking if ML should retrain from Excel data…")
                excel_retrain_from_feedback()
        except: pass


def morning_briefing(gc, vix_i, news, banks, events):
    print(f"\n{'█'*80}")
    print(f"☀️   MORNING BRIEFING  —  {datetime.now(IST).strftime('%A, %d %B %Y')}")
    print(f"{'█'*80}")
    print(f"    Prepared at {datetime.now(IST).strftime('%H:%M IST')}  |  Market opens 9:15 AM")

    # 1. High-impact events
    if events:
        print(f"\n🚨  HIGH-IMPACT EVENTS TODAY/TOMORROW:")
        for timing, desc in events:
            print(f"    {timing}: {desc}")
        if any("TODAY" in e[0] for e in events):
            print(f"\n    ⛔  MAJOR EVENT DAY — Consider reducing position size by 50%")
            print(f"    ⛔  Avoid holding positions through the event announcement")
    else:
        print(f"\n✅  No major scheduled events today — clean trading day")

    # 2. Global overnight
    print(f"\n🌍  OVERNIGHT GLOBAL MARKETS:")
    print(f"    Overall bias: {gc['bias']} (composite score: {gc['composite']:+.3f})")
    bull_mkts=[d for d in gc["cues"].values() if d["change"]>0 and d["price"]>0]
    bear_mkts=[d for d in gc["cues"].values() if d["change"]<0 and d["price"]>0]
    if bull_mkts: print("    UP  : " + ", ".join([m["name"]+" "+str(round(m["change"],1))+"%" for m in bull_mkts]))
    if bear_mkts: print("    DOWN: " + ", ".join([m["name"]+" "+str(round(m["change"],1))+"%" for m in bear_mkts]))

    # 3. India VIX
    vl=vix_i["level"]; vc=vix_i["change"]
    vix_interp = ("VERY HIGH — Options expensive, sell premium strategies better" if vl>22
                  else "HIGH — Larger moves expected today" if vl>18
                  else "NORMAL — Standard strategies work" if vl>14
                  else "LOW — Options cheap, buying strategies better")
    print(f"\n📊  INDIA VIX: {vl} ({vc:+.1f}%) → {vix_interp}")

    # 4. News sentiment
    print(f"\n📰  NEWS SENTIMENT: {news['bias']} (score: {news['score']})")
    print(f"    {news['bull_count']} bullish | {news['bear_count']} bearish | {news['total_found']} relevant headlines")
    if news["headlines"]:
        print(f"    Top headlines:")
        for h in news["headlines"][:5]:
            print(f"    • {h[:100]}")

    # 5. Bank sector health
    print(f"\n🏦  BANKING SECTOR: {banks['sector_bias']}")
    print(f"    {banks['n_bull']} of 12 banks bullish | Weighted breadth: {banks['w_breadth']}%")
    print(f"    Heavyweights: HDFC/ICICI/Kotak = {banks['heavy_bull']} bull / {banks['heavy_bear']} bear")

    # 6. Day bias
    bull_score = (1 if gc["composite"]>0.1 else -1 if gc["composite"]<-0.1 else 0) + \
                 (1 if vc<-1 else -1 if vc>1 else 0) + \
                 (1 if news["score"]>10 else -1 if news["score"]<-10 else 0) + \
                 (1 if banks["w_breadth"]>55 else -1 if banks["w_breadth"]<45 else 0)

    if bull_score >= 3:     day_bias="🟢 STRONG BULL BIAS — Favour BUY trades all day"
    elif bull_score >= 1:   day_bias="🟢 MILD BULL BIAS — BUY on dips preferred"
    elif bull_score <= -3:  day_bias="🔴 STRONG BEAR BIAS — Favour SELL trades all day"
    elif bull_score <= -1:  day_bias="🔴 MILD BEAR BIAS — SELL on bounces preferred"
    else:                   day_bias="⚪ NEUTRAL — Wait for ORB confirmation"

    day_log.morning_bias = day_bias

    print(f"\n{'═'*80}")
    print(f"📅  TODAY'S OVERALL BIAS: {day_bias}")
    print(f"{'═'*80}")

    # 7. Day plan
    print(f"\n📋  YOUR GAME PLAN FOR TODAY:")
    print(f"    8:30–9:15  Pre-market  : Monitor this briefing ✓")
    print(f"    9:15–9:45  Opening     : OBSERVE ONLY — track OR high/low, do NOT trade")
    print(f"    9:45–10:30 Post-OR     : FIRST TRADE WINDOW — best ORB setups")
    print(f"    10:30–12:30 Regular   : Trade with trend, use 15-min signals")
    print(f"    12:30–1:30 Lunch       : AVOID — low liquidity, avoid new entries")
    print(f"    1:30–2:30  Afternoon  : Resume with tighter stops")
    print(f"    2:30–3:15  Power Hour : Book profits aggressively, trail stops")
    print(f"    3:15 PM    MANDATORY EXIT: Close ALL open positions — no exceptions")

    print(f"\n    Capital    : Rs.{CAPITAL:,}")
    print(f"    Max risk   : Rs.{round(CAPITAL*RISK_PER_TRADE/100):,} per trade ({RISK_PER_TRADE}%)")
    print(f"    Max trades : {MAX_TRADES_DAY} trades max today")
    print(f"    Min R:R    : 1:{MIN_RR} -- skip every trade below this")
    print(f"    Max premium: Rs.{MAX_PREMIUM_PER_LOT} per lot (capital protection)")
    if PAPER_TRADE_MODE:
        print("\n    MODE: PAPER TRADING -- zero real money at risk")
        print("        Every signal is for practice only.")
        print("        Write each signal in a notebook. Track outcome at 3:15 PM.")
        print("        Go live only when paper win rate > 55% across 60 days.")
    else:
        print("\n    MODE: LIVE TRADING -- real money at risk. Stay disciplined.")
    if BEGINNER_MODE:
        print("\n    BEGINNER RULES -- Read every morning:")
        print("        1. NEVER trade 9:15-9:45 AM (Opening Range). Always wait.")
        print("        2. NEVER trade on Thursday (expiry day). Too dangerous.")
        print("        3. Set stop loss the MOMENT you enter. Never skip.")
        print("        4. NEVER move SL deeper when losing. Take the loss.")
        print(f"        5. After {MAX_TRADES_DAY} trades today -- STOP. No more. No exceptions.")
        print("        6. If you lose 2 trades in a row -- STOP for the day. No revenge.")
        print(f"        7. Only trade when confluence score is {MIN_CONFLUENCE}+.")
        print("        8. Exit ALL positions by 3:15 PM. Zero F&O overnight.")
    print("=" * 80)
    print("")

    session.morning_done = True
    # Send morning briefing to phone
    if ALERT_MORNING_BRIEF:
        alert_morning_briefing(
            day_bias  = day_log.morning_bias or "NEUTRAL",
            global_bias = gc["bias"],
            vix_level   = vix_i["level"],
            vix_chg     = vix_i["change"],
            news_bias   = news["bias"],
            bank_breadth= banks["w_breadth"],
            events      = events,
        )

# ══════════════════════════════════════════════════════════════════════════════
# TRADE GUIDANCE  (the complete per-signal output)
# ══════════════════════════════════════════════════════════════════════════════

def show_trade_guidance(ml_sig, conf, sess, oc, sess_type, patterns, regime, vix_i):
    spot=ml_sig["current"]; d=conf["direction"]; sc=conf["score"]
    vix_v=ml_sig.get("vix",15); atr_v=ml_sig.get("atr",300)
    now_str=datetime.now(IST).strftime("%H:%M")

    print(f"\n{'═'*80}")
    print(f"🎯  CONFLUENCE: {sc}/100  |  {'✅ TRADE SIGNAL' if conf['tradeable'] else '❌ NO TRADE'}")
    print(f"    BULL [{'█'*conf['bull_pts']:<18}] {conf['bull_pts']} pts")
    print(f"    BEAR [{'█'*conf['bear_pts']:<18}] {conf['bear_pts']} pts")
    print(f"\n  {'Indicator':<22} {'Signal':<28} Detail")
    print(f"  {'─'*22} {'─'*28} {'─'*24}")
    for name,sig,det in conf["checks"]:
        print(f"  {name:<22} {sig:<28} {det}")

    # ── NO TRADE GUIDANCE ────────────────────────────────────────────────────
    if not conf["tradeable"]:
        print(f"\n{'═'*80}")
        print(f"⛔  NO TRADE AT {now_str}")
        reasons=[]
        if ml_sig.get("adx",0)<MIN_ADX:
            reasons.append(f"Market is RANGING (ADX={ml_sig.get('adx',0):.0f} < {MIN_ADX}) — options decay in sideways markets")
        if ml_sig.get("confidence",0)<MIN_CONFIDENCE:
            reasons.append(f"ML confidence low ({ml_sig.get('confidence',0)}% < {MIN_CONFIDENCE}%) — uncertain direction")
        if sc<MIN_CONFLUENCE:
            reasons.append(f"Confluence low ({sc}/100 < {MIN_CONFLUENCE}) — indicators conflicting")
        if sess.trades_today>=MAX_TRADES_DAY:
            reasons.append(f"Daily trade limit reached ({sess.trades_today}/{MAX_TRADES_DAY})")
        if sess_type=="OPENING_RANGE":
            reasons.append("Still in opening range (9:15–9:45) — NEVER trade during this window")
        if sess_type=="LUNCH":
            reasons.append("Lunch period — low liquidity, wide spreads, avoid")

        for r in reasons:
            print(f"    • {r}")
        print(f"\n  ⏳  WHAT TO DO: Watch and wait. Next check in {CHECK_INTERVAL//60} minutes.")
        if sess_type=="OPENING_RANGE" and sess.or_high and sess.or_low:
            print(f"  📊  Current OR: High ₹{sess.or_high:,.2f} | Low ₹{sess.or_low:,.2f} | Range ₹{sess.or_high-sess.or_low:,.2f}")
            print(f"  ➡️   Watch for breakout above ₹{sess.or_high:,.2f} or below ₹{sess.or_low:,.2f} after 9:45 AM")
        return

    # ── TRADE PLAN ───────────────────────────────────────────────────────────
    atm_=atm(spot)

    # Strike selection based on score and regime
    if sc>=82:   off=2*STRIKE_INTERVAL; sz_label="AGGRESSIVE"
    elif sc>=72: off=STRIKE_INTERVAL;   sz_label="NORMAL"
    else:        off=0;                 sz_label="CONSERVATIVE (ATM)"

    # In ranging market, always trade ATM
    if regime=="RANGING": off=0; sz_label="ATM (ranging market)"

    rec=atm_+off if d=="BUY" else atm_-off
    typ="CE" if d=="BUY" else "PE"

    # Get real premium or estimate
    real_ep=get_real_premium(rec,typ,oc)
    ep=real_ep if real_ep and real_ep>0 else premium_est(spot,rec,typ,vix_v)
    ep_src="LIVE NSE ✅" if real_ep and real_ep>0 else "estimated"

    # ATR-based targets
    t1s=spot+atr_v*0.5  if d=="BUY" else spot-atr_v*0.5
    t2s=spot+atr_v*1.0  if d=="BUY" else spot-atr_v*1.0
    t3s=spot+atr_v*1.5  if d=="BUY" else spot-atr_v*1.5
    sls=spot-atr_v*0.5  if d=="BUY" else spot+atr_v*0.5

    t1p=premium_est(t1s,rec,typ,vix_v); t2p=premium_est(t2s,rec,typ,vix_v)
    t3p=premium_est(t3s,rec,typ,vix_v); slp=premium_est(sls,rec,typ,vix_v)

    # Risk and position sizing
    risk_per_lot=abs(ep-slp)*LOT_SIZE
    max_risk=CAPITAL*RISK_PER_TRADE/100
    lots=max(1,min(int(max_risk/risk_per_lot) if risk_per_lot>0 else 1, 5))

    p1=round((t1p-ep)*lots*LOT_SIZE); p2=round((t2p-ep)*lots*LOT_SIZE)
    p3=round((t3p-ep)*lots*LOT_SIZE); ls_=round((slp-ep)*lots*LOT_SIZE)
    avg_p=(p1+p2+p3)/3; rr=abs(avg_p/ls_) if ls_ else 0

    # Holding period
    if sess_type in ["POWER_HOUR","FORCE_EXIT"]:
        holding="⚡ SCALP only — exit before 3:15 PM mandatory"
    elif sc>=78 and ml_sig.get("adx",0)>=25:
        holding="📈 Full intraday — hold to T2/T3, trail stop after T1"
    elif sc>=68:
        holding="📊 Regular intraday — target T1 and T2, close by 3:00 PM"
    else:
        holding="⚡ Scalp only — target T1 only, quick in and out"

    # ── BEGINNER SAFETY FILTERS ─────────────────────────────────────────────
    if BEGINNER_MODE:
        if ep > MAX_PREMIUM_PER_LOT:
            print(f"\n  🔴  SIGNAL BLOCKED — Premium ₹{ep:.0f} exceeds your limit of ₹{MAX_PREMIUM_PER_LOT}")
            print(f"      With ₹{CAPITAL:,} capital, buying options above ₹{MAX_PREMIUM_PER_LOT} is too risky.")
            print(f"      Wait for a setup with cheaper premium (closer to expiry or OTM).\n")
            return
        if AVOID_EXPIRY_DAY and datetime.now(IST).weekday() == 3:
            print(f"\n  🔴  SIGNAL BLOCKED — Thursday Expiry Day (Beginner Mode)")
            print(f"      Expiry day has extreme time decay and unpredictable swings.")
            print(f"      OBSERVE ONLY today — no paper or real trades on Thursday.\n")
            return

    # ── R:R Check ────────────────────────────────────────────────────────────
    if rr < MIN_RR:
        print(f"\n  ⚠️  TRADE SKIPPED — R:R is only 1:{rr:.1f} (need 1:{MIN_RR})")
        print(f"      This trade is technically valid but the reward doesn't justify the risk.")
        print(f"      Wait for a better setup.\n")
        return

    # Paper trade vs live trade banner
    if PAPER_TRADE_MODE:
        print(f"\n{'='*80}")
        print(f"📝  PAPER TRADE SIGNAL — {now_str}  (Simulation — NO real money)")
        print(f"    {'🟢 BULLISH CALL (CE)' if d=='BUY' else '🔴 BEARISH PUT (PE)'}")
        print(f"    ► Write in notebook: {d} BANKNIFTY {rec} {typ} @ ₹{ep:.2f}")
        print(f"    ► Track outcome at 3:15 PM | Note WIN/LOSS | Build 60-day record")
        print(f"    Confluence: {sc}/100 | ML: {ml_sig['confidence']}% | Regime: {regime}")
        print(f"{'='*80}")
    else:
        print(f"\n{'█'*80}")
        print(f"✅  LIVE TRADE SIGNAL — {now_str}")
        print(f"    {'🟢 BUY CALL (CE) — Bullish trade' if d=='BUY' else '🔴 BUY PUT (PE) — Bearish trade'}")
        print(f"    Confluence: {sc}/100 | ML: {ml_sig['confidence']}% | Regime: {regime}")
        print(f"{'█'*80}")

    print(f"\n📋  CONTRACT DETAILS")
    print(f"    ┌─ Instrument  : BANKNIFTY {rec} {typ} (Expiry: {oc.get('expiry','nearest')})")
    print(f"    ├─ Entry Price : ₹{ep:.2f}  [{ep_src}]")
    print(f"    ├─ Position    : {lots} lot{'s' if lots>1 else ''} ({lots*LOT_SIZE} qty) — {sz_label}")
    print(f"    └─ Investment  : ₹{round(ep*lots*LOT_SIZE):,}")

    print(f"\n💰  PROFIT TARGETS")
    print(f"    ┌─ T1: Spot ₹{t1s:,.0f} → Premium ₹{t1p:.2f} → 📈 +₹{p1:,}  EXIT 30% here")
    print(f"    ├─ T2: Spot ₹{t2s:,.0f} → Premium ₹{t2p:.2f} → 📈 +₹{p2:,}  EXIT 40% here")
    print(f"    └─ T3: Spot ₹{t3s:,.0f} → Premium ₹{t3p:.2f} → 📈 +₹{p3:,}  EXIT 30% here")

    print(f"\n🛑  STOP LOSS  (NON-NEGOTIABLE)")
    print(f"    └─ SL: Spot ₹{sls:,.0f} → Premium ₹{slp:.2f} → 📉 -₹{abs(ls_):,}")
    print(f"       Set SL order IMMEDIATELY after entry. Do not move SL lower/higher in loss.")

    print(f"\n⚖️   RISK-REWARD: 1:{rr:.2f}  {'✅ Good setup' if rr>=2 else '✅ Acceptable' if rr>=1.5 else '⚠️ Marginal'}")
    print(f"⏱️   HOLDING: {holding}")

    if sess.vwap:
        vd=((spot-sess.vwap)/sess.vwap)*100
        print(f"📊  VWAP: ₹{sess.vwap:,.2f}  ({vd:+.2f}% — {'above' if vd>=0 else 'below'})")

    # OI levels
    if oc["success"]:
        print(f"\n📊  KEY OI LEVELS (from NSE):")
        if oc["resistance"]: print(f"    🔴 Resistance (heavy CE): {' | '.join([f'₹{r:,}' for r in oc['resistance']])}")
        if oc["support"]:    print(f"    🟢 Support (heavy PE)   : {' | '.join([f'₹{s_:,}' for s_ in oc['support']])}")
        if oc["max_pain"]:   print(f"    🎯 Max Pain strike       : ₹{oc['max_pain']:,}  ({oc['max_pain_dist']:+.1f}% from spot)")

    # Patterns contributing
    if patterns:
        bp_=[p for p in patterns if p["sig"]==d]
        if bp_:
            print(f"\n🕯️  CONFIRMING PATTERNS:")
            for p in bp_[:3]:
                print(f"    ✓ {p['name']} — {p['desc']}")

    # Groww steps
    print(f"\n📱  HOW TO PLACE ON GROWW:")
    print(f"    1. Open Groww → tap F&O → tap Bank Nifty")
    print(f"    2. Search: BANKNIFTY {rec} {typ}")
    print(f"    3. Tap BUY → enter qty {lots*LOT_SIZE} → place at MARKET or limit ₹{ep:.2f}")
    print(f"    4. Immediately place SL order: trigger ₹{slp+2:.2f}, limit ₹{slp-2:.2f}")
    print(f"    5. Set price alert at ₹{t1p:.2f} (T1) on Groww")
    print(f"    6. After T1 hit → MOVE SL to ₹{ep:.2f} (entry) → trade is now risk-free")

    # Log the signal
    # Write to Excel tracker
    _excel_row = excel_log_signal(
        date_str  = datetime.now(IST).strftime("%d-%m-%Y"),
        time_str  = now_str,
        direction = d, strike=rec, typ=typ,
        entry=ep, sl_prem=slp, t1=t1p, t2=t2p, t3=t3p,
        spot=spot, sl_spot=sls, lots=lots,
        score=sc, regime=regime,
        rsi   = ml_sig.get("rsi",50),
        adx   = ml_sig.get("adx",20),
        stoch_k=ml_sig.get("stoch_k",50),
        macd_h = ml_sig.get("macd_hist",0),
        bb_pctb= ml_sig.get("bb_pctb",0.5),
        vix    = ml_sig.get("vix",15),
        vix_chg= vix_i.get("change",0),
        sp500  = gc_data.get("composite",0) if "gc_data" in dir() else 0,
        news_score=news_data.get("score",0) if "news_data" in dir() else 0,
    )
    day_log.excel_signal_row = _excel_row

    day_log.log_signal(now_str, d, sc, ml_sig["confidence"], spot,
                       rec, typ, ep, slp, t1p, t2p, t3p)
    session.trades_today += 1
    print(f"\n    📋 Signal #{session.trades_today} of {MAX_TRADES_DAY} logged for today")
    # Send trade signal alert to phone
    alert_trade_signal(
        direction = d, strike = rec, typ = typ,
        entry = ep, sl = slp, t1 = t1p, t2 = t2p, t3 = t3p,
        score = sc, spot = spot, regime = regime,
        lots  = lots, rr = rr,
        expiry = oc.get("expiry","nearest weekly"),
        mode   = PAPER_TRADE_MODE,
    )
    # SL monitoring — check if open trade's SL is near current price
    day_log.active_trade = {
        "contract": f"BANKNIFTY {rec} {typ}",
        "direction": d, "entry": ep, "sl_spot": sls,
        "sl_prem": slp, "t1_spot": t1s, "t2_spot": t2s, "t3_spot": t3s,
        "t1_prem": t1p, "t2_prem": t2p, "t3_prem": t3p,
        "lots": lots, "targets_hit": [],
    }
    print(f"{'█'*80}")

# ══════════════════════════════════════════════════════════════════════════════
# FORCE EXIT ALERT
# ══════════════════════════════════════════════════════════════════════════════

def show_force_exit(spot, oc):
    print(f"\n{'█'*80}")
    print(f"🔔  3:15 PM — MANDATORY EXIT ALERT")
    print(f"{'█'*80}")
    print(f"    Bank Nifty Spot: ₹{spot:,.2f}")
    print(f"    TIME IS UP — Close ALL open Bank Nifty F&O positions NOW")
    print(f"\n    WHY: Options time decay accelerates massively in last 15 minutes.")
    print(f"    Holding past 3:15 PM on expiry day can cause rapid premium collapse.")
    print(f"\n    HOW TO EXIT on Groww:")
    print(f"    1. F&O → Positions → find your open BANKNIFTY position")
    print(f"    2. Tap → Select 'Close position' → Market order")
    print(f"    3. Confirm — done")
    print(f"\n    NO EXCEPTIONS. Even if the trade is running at a loss.")
    print(f"    Taking a small loss today is better than a catastrophic loss.")
    print(f"{'█'*80}")
    # Send force exit alert to phone
    alert_force_exit(spot_tmp)
    day_log.summary()

# ══════════════════════════════════════════════════════════════════════════════
# TRAINING PHASE
# ══════════════════════════════════════════════════════════════════════════════

print("\n" + "═"*80)
print("⏳  PHASE 1/4 — Fetching news & events…")
print("═"*80)
news_data = fetch_news_sentiment()
events    = fetch_rbi_events()
print(f"   ✓ News: {news_data['bias']} | {news_data['bull_count']} bull / {news_data['bear_count']} bear headlines")
if events:
    for e in events: print(f"   ⚠️  Event: {e[0]} — {e[1]}")

print("\n" + "═"*80)
print("⏳  PHASE 2/4 — Downloading 2-year historical data + engineering features…")
print("═"*80)
hist_raw  = yf.download(TICKER, period="2y", auto_adjust=True, progress=False)
print(f"   ✓ {len(hist_raw)} daily bars downloaded")
hist_feat = build_features(hist_raw, news_score=news_data["score"])
print(f"   ✓ {len(ALL_FEATURES)} features engineered")

print("\n" + "═"*80)
print("⏳  PHASE 3/4 — HMM Regime Detection…")
print("═"*80)
regime_det.fit(hist_feat.dropna())
hf2 = hist_feat.dropna().copy()
hf2["regime"] = regime_det.model.predict(
    np.hstack([hf2["Ret_1d"].values.reshape(-1,1),
               hf2["Vol_10"].values.reshape(-1,1)]))
hf2["regime_lbl"] = hf2["regime"].map(regime_det.regime_map)
tr_d = hf2[hf2["regime_lbl"]=="TRENDING"]
rg_d = hf2[hf2["regime_lbl"]=="RANGING"]
print(f"   ✓ Trending bars: {len(tr_d)} | Ranging bars: {len(rg_d)}")

print("\n" + "═"*80)
print("⏳  PHASE 4/4 — Training ML ensemble (LightGBM + XGBoost + LR)…")
print("═"*80)
print("\n  📈 TRENDING model:")
models["TRENDING"].train(tr_d if len(tr_d)>80 else hf2, ALL_FEATURES)
print("\n  📦 RANGING model:")
models["RANGING"].train(rg_d  if len(rg_d)>80  else hf2, ALL_FEATURES)
print("\n  🌐 ALL-REGIME fallback:")
model_all.train(hf2, ALL_FEATURES)

print("\n" + "=" * 80)
print("SYSTEM READY — All models trained")
print("=" * 80 + "\n")
test_alerts()
_init_excel_tracker()
_retrain_cycle_counter = [0]
_init_excel_tracker()
# Track retraining cycles
_retrain_cycle_counter = [0]

# ══════════════════════════════════════════════════════════════════════════════
# CACHE VARIABLES
# ══════════════════════════════════════════════════════════════════════════════
_bank_c=None; _bank_t=None; BANK_R=15
_glob_c=None; _glob_t=None; GLOB_R=10
_oc_c  =None; _oc_t  =None; OC_R=5
_news_c=None; _news_t=None; NEWS_R=60   # refresh news once per hour

# ══════════════════════════════════════════════════════════════════════════════
# MAIN LOOP
# ══════════════════════════════════════════════════════════════════════════════

print("=" * 80)
print("🚀  LIVE MONITORING STARTED  (Ctrl+C to stop)")
print(f"    Refreshing every {CHECK_INTERVAL} seconds")
print("=" * 80)

_crash_count = 0
_MAX_CRASHES = 10
sess_type = "PRE_MARKET"  # default before first loop

# ── Send startup Telegram message immediately ──────────────────────────────
import datetime as _dt_mod
_now_ist = _dt_mod.datetime.now(__import__("pytz").timezone("Asia/Kolkata"))
try:
    import requests as _rq
    if TELEGRAM_TOKEN and TELEGRAM_CHAT_ID:
        _msg = (
            "<b>✅ BankNifty Bot Started on GitHub</b>\n"
            f"Time: {_now_ist.strftime('%d %b %Y %H:%M IST')}\n"
            f"Mode: {'PAPER TRADE' if PAPER_TRADE_MODE else 'LIVE TRADE'}\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "Bot is running on GitHub servers\n"
            "Auto-schedule: Mon-Fri 8:15 AM IST\n"
            "You will get trade signals during market hours"
        )
        _rq.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            data={"chat_id": TELEGRAM_CHAT_ID, "text": _msg,
                  "parse_mode": "HTML"},
            timeout=10
        )
        print("Startup Telegram message sent!")
    else:
        print("No Telegram credentials found.")
except Exception as _e:
    print(f"Startup message failed: {_e}")

while True:
    try:
        # Skip weekends and non-market hours automatically
        if not is_active_session():
            _secs = seconds_until_market()
            _hrs  = int(_secs//3600)
            _mins = int((_secs%3600)//60)
            print(f"Outside market hours. Sleeping {_hrs}h {_mins}m...")
            import time as _tsleep
            _tsleep.sleep(min(_secs, 1800))   # check every 30 min max
            continue
    except Exception as _skip:
        now     = datetime.now(IST)
        today   = now.date()
        cur_t   = now.time()
        sess_type = session_name(cur_t)

        # ── Daily reset ────────────────────────────────────────────────────
        if session.last_date != today:
            print(f"\n🆕  NEW TRADING DAY: {today}")
            session.reset(); day_log.reset()
            session.last_date = today; day_log.date = str(today)
            # Refresh news at start of each day
            news_data = fetch_news_sentiment()
            events    = fetch_rbi_events()

        # ── Force exit alert ───────────────────────────────────────────────
        if sess_type == "FORCE_EXIT" and not session.force_exit_alerted:
            session.force_exit_alerted = True
            # Get spot price
            intra_tmp = yf.download(TICKER,period="1d",interval="1m",
                                    auto_adjust=True,progress=False)
            spot_tmp = float(safe_s(intra_tmp,"Close").iloc[-1]) if len(intra_tmp)>0 else 0
            oc_tmp = _oc_c if _oc_c else {"success":False}
            show_force_exit(spot_tmp, oc_tmp)
            time.sleep(CHECK_INTERVAL); continue

        if sess_type == "CLOSED":
            if not getattr(session,"closed_summary_done",False):
                session.closed_summary_done=True
                day_log.summary()
                # Write daily summary to Excel
                wins_ = 0
                excel_log_daily(
                    str(today),
                    len(day_log.signals),
                    session.trades_today,
                    0, 0,   # wins/losses updated from outcomes
                    0,      # P&L from journal
                    day_log.morning_bias or "NEUTRAL",
                )
                # Check if we should retrain ML from real trade data
                _retrain_cycle_counter[0] += 1
                if _retrain_cycle_counter[0] % 5 == 0:   # check every 5 days
                    result = excel_retrain_from_feedback()
                    if result:
                        print(f"  🧠 ML retrained from real trades!")
                print(f"\n  Market closed. See you tomorrow! 🌙")
                print(f"  📊 Excel journal updated: {EXCEL_PATH}")
                log.info(f"Excel updated: {EXCEL_PATH}")
            time.sleep(300); continue  # check every 5 min after close

        # ── Morning briefing ───────────────────────────────────────────────
        if not session.morning_done:
            b_c=None; b_t=None
            if _bank_c is None:
                print("  🏦  Fetching bank data for morning briefing…")
                _bank_c=analyse_banks(); _bank_t=now
            if _glob_c is None:
                print("  🌍  Fetching global data…")
                gc_=fetch_global(); vx_=fetch_vix()
                _glob_c=(gc_,vx_); _glob_t=now
            gc_data,vix_data=_glob_c
            morning_briefing(gc_data, vix_data, news_data, _bank_c, events)

        print(f"\n{'═'*80}")
        print(f"⏰  {now.strftime('%H:%M:%S  IST')}  |  Session: {sess_type.replace('_',' ')}")
        print(f"    Trades today: {session.trades_today}/{MAX_TRADES_DAY}")
        print(f"{'═'*80}")

        # ── Refresh caches ─────────────────────────────────────────────────
        oc_mins   = 999 if _oc_t   is None else (now-_oc_t).seconds//60
        b_mins    = 999 if _bank_t is None else (now-_bank_t).seconds//60
        g_mins    = 999 if _glob_t is None else (now-_glob_t).seconds//60
        news_mins = 999 if _news_t is None else (now-_news_t).seconds//60

        if oc_mins>=OC_R:
            print("  📊  Fetching NSE options chain…")
            _oc_c=fetch_options_chain(); _oc_t=now
        if b_mins>=BANK_R:
            _bank_c=analyse_banks(); _bank_t=now
        if g_mins>=GLOB_R:
            gc_=fetch_global(); vx_=fetch_vix()
            _glob_c=(gc_,vx_); _glob_t=now
        if news_mins>=NEWS_R:
            news_data=fetch_news_sentiment(); _news_t=now

        oc=_oc_c or {"success":False}
        gc_data,vix_data=_glob_c
        bd=_bank_c

        # ── Live spot price ────────────────────────────────────────────────
        spot = oc["spot"] if oc.get("success") and oc["spot"]>0 else 0
        intra = yf.download(TICKER,period="5d",interval=INTRADAY_TF,
                            auto_adjust=True,progress=False)
        today_15m = intra[intra.index.date==today]

        if len(today_15m)>0:
            px15=float(safe_s(today_15m,"Close").iloc[-1])
            if spot==0: spot=px15
            try: session.vwap=vwap_f(today_15m)
            except: pass
            if sess_type!="OPENING_RANGE" and not session.or_set and len(today_15m)>=2:
                session.set_or(float(safe_s(today_15m.iloc[:2],"High").max()),
                               float(safe_s(today_15m.iloc[:2],"Low").min()))
            if session.or_set: session.check_orb(spot)

        patterns = find_patterns(today_15m) if len(today_15m)>0 else []

        # ── Display key numbers ────────────────────────────────────────────
        if spot>0:
            print(f"\n💰  BANK NIFTY SPOT: Rs.{spot:,.2f}")
            # Log price to Excel if trade is open
            if getattr(day_log,"excel_signal_row",None) and                getattr(day_log,"active_trade",None):
                at_ = day_log.active_trade
                excel_log_price(
                    datetime.now(IST).strftime("%d-%m-%Y"),
                    datetime.now(IST).strftime("%H:%M"),
                    spot,
                    at_.get("entry_spot",spot),
                    at_.get("t1_spot",spot),
                    at_.get("sl_spot",spot),
                )
            # ── Active trade monitoring ───────────────────────────────────
            at = getattr(day_log, "active_trade", None)
            if at:
                d_at   = at["direction"]
                c_at   = at["contract"]
                sl_sp  = at["sl_spot"]
                # Check SL proximity (within 0.3%)
                sl_dist = abs(spot - sl_sp) / sl_sp * 100
                if sl_dist < 0.3 and "SL_warned" not in at.get("targets_hit", []):
                    print(f"  ⚠️  SL WARNING: Spot Rs.{spot:,.0f} near SL Rs.{sl_sp:,.0f}")
                    alert_sl_warning(spot, sl_sp, c_at)
                    at["targets_hit"].append("SL_warned")
                # Check targets hit
                for tnum, tkey in [(1,"t1_spot"),(2,"t2_spot"),(3,"t3_spot")]:
                    if tnum not in at.get("targets_hit",[]):
                        t_sp   = at[tkey]
                        t_prem = at[f"t{tnum}_prem"]
                        profit = round((t_prem - at["entry"]) * at["lots"] * 25)
                        hit = (spot >= t_sp if d_at=="BUY" else spot <= t_sp)
                        if hit:
                            print(f"  🎯  TARGET {tnum} HIT! Spot Rs.{spot:,.0f} >= Rs.{t_sp:,.0f}")
                            alert_target_hit(tnum, spot, t_prem, profit, c_at)
                            at["targets_hit"].append(tnum)
                            excel_update_outcome(
                                day_log.excel_signal_row,
                                f"T{tnum} HIT",
                                t_prem,
                                profit
                            )
                            if tnum == 3:
                                at = None  # all targets hit, clear trade
                                day_log.active_trade = None
        if session.or_set:
            print(f"📊  OPENING RANGE: High ₹{session.or_high:,.2f} | Low ₹{session.or_low:,.2f} | Range ₹{session.or_high-session.or_low:,.0f}")
            if session.orb_det:
                print(f"    {'🟢 ORB BREAKOUT — BULLISH' if session.orb_dir=='BULLISH' else '🔴 ORB BREAKDOWN — BEARISH'} confirmed!")
        if session.vwap:
            vd=((spot-session.vwap)/session.vwap)*100 if spot>0 else 0
            print(f"📈  VWAP: ₹{session.vwap:,.2f}  ({vd:+.2f}%)")

        # PCR + Max Pain
        if oc["success"]:
            print(f"⚖️   PCR: {oc['pcr']}  →  {oc['pcr_signal']}")
            print(f"🎯  Max Pain: ₹{oc['max_pain']:,}  ({oc['max_pain_dist']:+.1f}% from spot)")

        # Global + VIX
        print(f"🌍  Global: {gc_data['bias']}  |  VIX: {vix_data['level']} ({vix_data['change']:+.1f}%) {vix_data['signal']}")
        print(f"🏦  Banks: {bd['n_bull']} Bull / {bd['n_bear']} Bear | Sector: {bd['sector_bias']}")
        print(f"📰  News: {news_data['bias']}")

        # Patterns
        if patterns:
            print(f"\n🕯️  15-MIN PATTERNS:")
            for p in patterns:
                e="🟢" if p["sig"]=="BUY" else "🔴" if p["sig"]=="SELL" else "⚪"
                print(f"    {e} {p['name']} [{p['conf']}%] — {p['desc']}")

        # ── ML Signal ─────────────────────────────────────────────────────
        daily_raw  = yf.download(TICKER,period="1y",auto_adjust=True,progress=False)
        daily_feat = build_features(daily_raw, news_score=news_data["score"])
        regime     = regime_det.predict(daily_feat.dropna())
        active     = models.get(regime,model_all)
        if active.lgbm is None: active=model_all
        ml_sig     = active.predict(daily_feat.dropna())
        if ml_sig is None: ml_sig=model_all.predict(daily_feat.dropna())

        df_ = daily_feat.dropna()
        for fld,col in [("rsi","RSI"),("adx","ADX"),("di_diff","DI_diff"),
                        ("st_dir","ST_dir"),("bb_pctb","BB_pctb"),
                        ("stoch_k","Stoch_K"),("cci","CCI"),
                        ("macd_hist","MACD_hist"),("vix","VIX"),
                        ("vol_ratio","Vol_R10"),("atr","ATR")]:
            try: ml_sig[fld]=round(float(df_[col].iloc[-1]),2)
            except: ml_sig[fld]=0

        if spot>0: ml_sig["current"]=spot

        print(f"\n🧠  Regime: {regime} | ML: {ml_sig['direction']} "
              f"{ml_sig['confidence']}% ({ml_sig['strength']}) | CV: {ml_sig['cv_accuracy']}%")

        # ── Session-specific guidance ──────────────────────────────────────
        if sess_type == "OPENING_RANGE":
            print(f"\n{'═'*80}")
            print(f"🔥  OPENING RANGE IN PROGRESS (9:15 – 9:45 AM)")
            print(f"    DO NOT TRADE — just observe and track the range")
            print(f"    Watching for: ORB breakout above ₹{session.or_high:,.2f}" if session.or_set else "    Range not set yet — wait")
            print(f"{'═'*80}")

        elif sess_type == "LUNCH":
            print(f"\n{'═'*80}")
            print(f"🍽️   LUNCH BREAK (12:30 – 1:30 PM)")
            print(f"    Low liquidity — wide bid-ask spreads — avoid new positions")
            print(f"    If you have open positions: trail your stop, don't add")
            print(f"{'═'*80}")

        elif sess_type == "POWER_HOUR":
            print(f"\n⚡  POWER HOUR — 2:30 PM onwards")
            print(f"    • Exit 50% of remaining positions NOW")
            print(f"    • Trail stop on remaining 50%")
            print(f"    • No NEW trades after 3:00 PM")
            print(f"    • MANDATORY FULL EXIT by 3:15 PM\n")
            # Still show confluence but with warning
            conf = calc_confluence(ml_sig,session,bd,patterns,gc_data,vix_data,oc,regime,news_data)
            show_trade_guidance(ml_sig,conf,session,oc,sess_type,patterns,regime,vix_data)

        else:
            # Regular trading session
            conf = calc_confluence(ml_sig,session,bd,patterns,gc_data,vix_data,oc,regime,news_data)
            show_trade_guidance(ml_sig,conf,session,oc,sess_type,patterns,regime,vix_data)

    except KeyboardInterrupt:
        print("\nStopped by user.")
        log.info("Server stopped by user (KeyboardInterrupt)")
        try:
            os.remove(PID_FILE)
        except Exception:
            pass
        break
    except Exception as e:
        import traceback
        _crash_count += 1
        _tb = traceback.format_exc()
        log.error(f"CRASH #{_crash_count}: {e}")
        log.error(_tb)
        print(f"\nERROR #{_crash_count}: {e}")
        if _crash_count <= _MAX_CRASHES:
            _send_tg_crash(TELEGRAM_TOKEN, TELEGRAM_CHAT_ID, e)
            print(f"Auto-restarting in 60 seconds... ({_crash_count}/{_MAX_CRASHES})")
            import time as _tsleep
            _tsleep.sleep(60)
        else:
            print(f"Too many crashes ({_MAX_CRASHES}). Stopping server.")
            log.error("Max crashes reached. Server stopped.")
            break

    next_t = CHECK_INTERVAL if sess_type not in ["CLOSED","PRE_MARKET","OUTSIDE_HOURS"] else 300
    print(f"\n💤  Next refresh in {next_t//60} min {next_t%60} sec…  (Ctrl+C to stop)\n")
    time.sleep(next_t)
